# -*- coding: utf-8 -*-
"""现成的 Markdown 转换稿 → 待核记录 + 一份本地 Excel。

`gaz convert`(转换交给 zhiconv)是给扫描件预备的。手里若已经有转好的 .md ——
别处转的、自己抄的、从数字方志库拷出来的 —— 那一步不必走,直接从这里进。

要紧的差别有两处:

**一、编码。** Windows 上存下来的稿子可能是 GB18030、可能带 BOM、也可能是
UTF-16。挨个试,试通了记下来告诉你,别让一堆乱码悄悄流进表里。

**二、页码。** 转换稿多半不留页码。本工具认的是每页正文前的 `<!-- p.123 -->`,
抽取时据以回注出处;别处转来的没有这一手。所以先认一遍常见的页码写法
(`第123页`、`- 123 -`、`[123]` 之类),认出来就归一成 `<!-- p.N -->`;
认不出也不要紧 —— 出处退到篇章节,「北京工业志·电子志·第三章」仍查得回去。
"""

import os
import re
from collections import Counter

from . import cndate
from . import extract as EX

ENCODINGS = ["utf-8-sig", "utf-8", "gb18030", "big5", "utf-16", "latin-1"]

# 统计块:字段名 ↔ 表头。写出去、读回来共用一份,免得两头对不上
STAT_COLS = [("staff", "职工总数"), ("tech", "技术人员"), ("plant", "厂房面积"),
             ("floor", "建筑面积"), ("assets", "固定资产"), ("output", "工业总产值"),
             ("sales", "销售收入"), ("profit", "实现利润")]

# 页码的常见写法。每条给出正则与取数的组号;认哪一条,看谁的数字最像页码。
PAGE_PATTERNS = [
    ("已是本工具的写法", re.compile(r"^[ \t]*<!--\s*p\.(\d{1,4})\s*-->[ \t]*$", re.M)),
    ("HTML 注释", re.compile(r"^[ \t]*<!--\s*(\d{1,4})\s*-->[ \t]*$", re.M)),
    ("第N页", re.compile(r"^[ \t]*第\s*(\d{1,4})\s*页[ \t]*$", re.M)),
    ("方括号", re.compile(r"^[ \t]*\[\[?\s*(\d{1,4})\s*\]?\][ \t]*$", re.M)),
    ("花括号", re.compile(r"^[ \t]*\{\{?\s*(\d{1,4})\s*\}?\}[ \t]*$", re.M)),
    ("破折号夹注", re.compile(r"^[ \t]*[-—–·\*]{1,2}\s*(\d{1,4})\s*[-—–·\*]{1,2}[ \t]*$", re.M)),
    ("锚点", re.compile(r'^[ \t]*<a\s+(?:id|name)="(?:page|p)?(\d{1,4})"\s*/?>(?:</a>)?[ \t]*$',
                        re.M | re.I)),
    ("P123", re.compile(r"^[ \t]*[Pp]\.?\s*(\d{1,4})[ \t]*$", re.M)),
    ("孤零数字", re.compile(r"^[ \t]*(\d{1,4})[ \t]*$", re.M)),
]


# 汉字(含中日韩标点、全角符号)之间的空格一律是转换留下的,没有意义
_CJK = "\u3400-\u4dbf\u4e00-\u9fff\u3000-\u303f\uff00-\uffef"
# 只吃空格,不吃制表符 —— 制表符多半是分栏的(订正表就是 TSV),吃掉就把两列并了
CJK_GAP = re.compile("(?<=[%s]) +(?=[%s])" % (_CJK, _CJK))


# 型号里的连字符被认成了汉字「一」:TQ一16、DJS一131、X一2型、JDK一331。
# 《上海电子仪表工业志》第一章一篇就 117 处。夹在拉丁字母与数字之间的「一」,
# 中文里没有这种写法,一律是连字符认岔了 —— 不改,型号认不出来,更要紧的是
# 跟总表里的 DJS-131 对不上,判重拦不住,同一台机器要收两遍。
DASH_ONE = re.compile(r"(?<=[A-Za-z])[一―−](?=[0-9])")


def fix_model_dash(text):
    """把型号里认成「一」的连字符改回来。返回 (改过的文本, 改了几处)。"""
    return DASH_ONE.subn("-", text)


def squeeze_cjk_spaces(text):
    """去掉汉字中间的空格。

    转换稿常在原书断行处留一个空格,成了「北京计算 机一厂」。空格是找厂名时
    往左走的边界,留着就截出「机一厂」这种鬼名字。标题和表格不动 —— 那里的
    空格是分栏用的。"""
    out = []
    for line in text.splitlines():
        s = line.lstrip()
        out.append(line if s[:1] in ("#", "|") or s.startswith("```")
                   else CJK_GAP.sub("", line))
    return "\n".join(out)


def read_text(path):
    """挨个试编码,返回 (正文, 用的哪一种)。顺手把换行和汉字间的空格归一。"""
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ENCODINGS:
        try:
            text = raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
        # 解出来若满是替换符或几乎没有汉字,多半是解错了,接着试
        if "�" in text[:4000]:
            continue
        if enc == "latin-1" and len(re.findall(r"[一-鿿]", text[:4000])) < 5:
            continue
        return squeeze_cjk_spaces(
            text.replace("\r\n", "\n").replace("\r", "\n")), enc
    return raw.decode("utf-8", errors="replace").replace("\r\n", "\n"), "utf-8(有乱码)"


def load_fixes(path):
    """读字形订正表:一行一条「错<TAB>对」,# 开头算注解。

    转换稿总有认错的字,「安徽无线电厂」成了「安做无线电厂」。这种错改不出
    规律,只能一本书一张表,跟着书走。"""
    if not path:
        return []
    text, _enc = read_text(path)
    out = []
    for line in text.splitlines():
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        cell = line.rstrip("\n").split("\t")
        if len(cell) >= 2 and cell[0]:
            out.append((cell[0], cell[1]))
    return out


def apply_fixes(text, fixes):
    """逐条改,并回报每条改了几处 —— 表里写错了字,得看得见,不能默默不改。"""
    ledger = []
    for bad, good in fixes:
        n = text.count(bad)
        if n:
            text = text.replace(bad, good)
        ledger.append((bad, good, n))
    return text, ledger


def _monotonic(nums):
    """页码该是大体递增的;脚注号 [1][2][1] 不是。返回递增的比例。"""
    if len(nums) < 3:
        return 0.0
    ups = sum(1 for a, b in zip(nums, nums[1:]) if b >= a)
    return ups / float(len(nums) - 1)


def detect_page_marks(text):
    """把各种页码写法都试一遍,按「像不像页码」排出来。"""
    out = []
    for name, pat in PAGE_PATTERNS:
        nums = [int(m.group(1)) for m in pat.finditer(text)]
        if len(nums) < 3:
            continue
        mono = _monotonic(nums)
        out.append({"name": name, "pattern": pat, "count": len(nums),
                    "monotonic": mono, "first": nums[0], "last": nums[-1],
                    "ok": mono >= 0.9})
    out.sort(key=lambda d: (d["ok"], d["count"]), reverse=True)
    return out


def normalize_page_marks(text, force=None):
    """把认出来的页码写法归一成 `<!-- p.N -->`。返回 (正文, 用了哪条, 改了几处)。"""
    cands = [c for c in detect_page_marks(text) if c["ok"]]
    if force:
        cands = [c for c in detect_page_marks(text) if c["name"] == force] or cands
    if not cands:
        return text, None, 0
    pick = cands[0]
    if pick["name"] == "已是本工具的写法":
        return text, pick["name"], pick["count"]
    new = pick["pattern"].sub(lambda m: "<!-- p.%s -->" % int(m.group(1)), text)
    return new, pick["name"], pick["count"]


# ---------------------------------------------------------------- 接断行

SENT_END = "。！？；.!?;：:」』】》）)"
_SKIP = re.compile(r"^\s*(?:#{1,6}\s|[-*+]\s|\d+[.、)]\s|>|\||```|<!--|---\s*$)")


def hard_wrapped(text):
    """看这份稿子是不是照原书的行宽硬断的。

    转换稿常把版心一行原样存成一行:「北京市半导体」「器件研究所」各一行,
    厂名就断成了两截,怎么认都认不出。判据是「多少行没收在句读上」。"""
    body = [l.strip() for l in text.splitlines()
            if l.strip() and not _SKIP.match(l)]
    if len(body) < 8:
        return False, 0.0
    open_ended = sum(1 for l in body if l[-1] not in SENT_END)
    share = open_ended / float(len(body))
    return share >= 0.3, share


def reflow_soft(text):
    """把硬断的行接回段落:上一行没收在句读上,下一行又不是标题/表格/列表,就接上。

    比 `gaz md` 的那一套轻:那边要对付 OCR 的碎行,这边只接明显断开的。
    已经一段一行的稿子跑一遍也不会有事 —— 每行都收在句号上,一行也不会接。"""
    out = []
    for line in text.splitlines():
        s = line.rstrip()
        if not s.strip() or _SKIP.match(s) or not out or not out[-1].strip():
            out.append(s)
            continue
        prev = out[-1].rstrip()
        if _SKIP.match(prev) or not prev:
            out.append(s)
            continue
        if prev[-1] in SENT_END:
            out.append(s)
        else:
            out[-1] = prev + s.lstrip()
    return "\n".join(out)


# ---------------------------------------------------------------- 看一眼

def inspect(text, path=""):
    """先看看这份稿子长什么样 —— 抽取之前值得花十秒钟。"""
    lines = text.splitlines()
    heads = [l for l in lines if re.match(r"^#{1,6}\s+\S", l)]
    levels = Counter(len(re.match(r"^(#+)", h).group(1)) for h in heads)
    tables = sum(1 for l in lines if l.strip().startswith("|"))

    wrapped, share = hard_wrapped(text)
    scan = reflow_soft(text) if wrapped else text
    units = Counter()
    for l in scan.splitlines():
        for nm in EX.unit_names(l):
            units[nm] += 1

    cues = {k: len(re.findall(v, text)) for k, v in [
        ("前身 / 原名", r"前身|原名|原为|原系"),
        ("改名 / 更名", EX.RENAME),
        ("划归 / 隶属", EX.TRANSFER),
        ("合并 / 并入", EX.MERGE),
        ("始建 / 成立", EX.BIRTH),
        ("撤销 / 停办", EX.DEATH),
        ("试制 / 投产", EX.MADE),
        ("职工 / 产值", r"职工|技术人员|总产值|销售收入|固定资产|利润"),
    ]}

    return {
        "path": path, "chars": len(text), "lines": len(lines),
        "headings": len(heads), "levels": dict(sorted(levels.items())),
        "sample_headings": [h.strip() for h in heads[:12]],
        "table_lines": tables, "wrapped": wrapped, "wrap_share": share,
        "pages": detect_page_marks(text),
        "units": units.most_common(15),
        "unit_total": len(units),
        "cues": cues,
    }


def report(info, log=print):
    log("稿子:%s" % (info["path"] or "(未命名)"))
    log("  %d 字,%d 行,%d 处标题%s" % (
        info["chars"], info["lines"], info["headings"],
        ("(层级 " + "、".join("%d 级×%d" % (k, v) for k, v in info["levels"].items()) + ")")
        if info["levels"] else ""))
    if info["sample_headings"]:
        log("  头几处标题:")
        for h in info["sample_headings"][:8]:
            log("     " + h[:60])
    if info["table_lines"]:
        log("  %d 行像表格 —— 表里的数字本工具不还原,那部分仍要手录" % info["table_lines"])
    if info["wrapped"]:
        log("  %.0f%% 的行没收在句读上 —— 照原书行宽硬断的,抽取前会先接回段落"
            % (info["wrap_share"] * 100))

    if info["pages"]:
        for c in info["pages"][:3]:
            log("  页码写法「%s」%d 处,%d→%d,递增率 %.0f%%%s"
                % (c["name"], c["count"], c["first"], c["last"], c["monotonic"] * 100,
                   "  ← 采用" if c["ok"] and c is info["pages"][0] else ""))
    else:
        log("  没认出页码 —— 出处会退到篇章节(如「…·第三章」),仍查得回去")

    log("  认出 %d 个单位名,最常出现的几个:" % info["unit_total"])
    for nm, n in info["units"][:8]:
        log("     %s ×%d" % (nm, n))

    log("  志书套语的出现次数(抽取全靠这些词):")
    for k, v in info["cues"].items():
        log("     %-12s %d" % (k, v))
    weak = [k for k, v in info["cues"].items() if v == 0]
    if weak:
        log("  注意:%s 一次都没出现 —— 这几类字段多半抽不到。" % "、".join(weak))


# ---------------------------------------------------------------- 本地 Excel

def write_xlsx(path, res, city="", book="", stats_year=1990, log=print):
    """把抽出来的东西写成一份本地工作簿,版式与 CN_Electronic_Industry.xlsx 一致,
    另附一张「待核」表,把出处、原文、置信一并摆上,好在 Excel 里逐条核对。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    stat_labels = [label for _k, label in STAT_COLS]
    stat_keys = [k for k, _ in STAT_COLS]

    def num(v):
        s = str(v if v is not None else "").strip()
        if s == "":
            return None
        if re.fullmatch(r"\d{8}", s):
            return int(s)
        try:
            f = float(s)
            return int(f) if f == int(f) else f
        except ValueError:
            return s

    # ---- 厂所名录:两行表头,与原表一模一样
    sheet = "Fact and Comp-" + (city or "Local")
    ws = wb.create_sheet(sheet)
    ws.append(["", "Industry", "Product", "Start Date", "End Date", "Founder", "City", "Add.",
               stats_year] + [""] * 7 + ["Remark", "Source", "别名"])
    ws.append([""] * 8 + stat_labels + ["", ""])
    for r in res["units"]:
        ws.append([r.get("Unit", ""), r.get("Industry", ""), r.get("Product", ""),
                   num(r.get("Start Date")), num(r.get("End Date")), r.get("Founder", ""),
                   r.get("City", city), r.get("Add.", "")]
                  + [num(r.get(k)) for k in stat_keys]
                  + [r.get("Remark", ""), r.get("Source", ""), r.get("别名", "")])
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=16)
    ws.cell(row=1, column=19).font = Font(bold=True)
    # 这张表照原表体例生成,好让你一眼看出将来落在地图上是什么样 —— 但读回来
    # 只读「待核」。不写明白,在这儿改半天不算数,还没有一处告诉你。
    note = ws.cell(row=1, column=21,
                   value="↑ 此表照 CN_Electronic_Industry.xlsx 的体例生成,供预览与粘贴。"
                         "改这里不算数 —— 要改请改「待核」表。")
    note.font = Font(italic=True, color="996600")
    ws.cell(row=1, column=9).alignment = Alignment(horizontal="center")
    for c in range(1, 19):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=2, column=c).font = Font(bold=True)
    ws.freeze_panes = "A3"

    def flat(name, cols, rows, keys=None):
        w = wb.create_sheet(name)
        w.append(["取否"] + cols)
        for c in range(1, len(cols) + 2):
            w.cell(row=1, column=c).font = Font(bold=True)
        for r in rows:
            w.append([""] + [num(r.get(k)) if k in ("Time", "From") else r.get(k, "")
                             for k in (keys or cols)])
        w.freeze_panes = "B2"
        return w

    flat("Semi-Product", ["Product", "别名", "Research Insti", "Factory", "产量", "Time",
                          "Personnel", "Remark"], res["semi"])
    flat("Comp-Product", ["Product", "字长", "内存", "Speed（次秒）", "Research Insti",
                          "Factory", "用户", "产量", "别名", "Time", "Personnel", "Remark"], res["comp"])
    # 待核那份的沿革表也照总表的次序摆:序、单位、名称、起、至、关系
    nh = wb.create_sheet("Name-History")
    # 表头写成中文,把话说死:「Unit」看着像「这一行这家单位叫什么」,而它其实
    # 是钥匙 —— 一家单位的几行都写同一个今名。改叫「单位(今名)」就不会看岔。
    nh.append(["取否", "序", "单位(今名)", "当时名称", "自哪年起", "Remark", "Source"])
    for c in range(1, 8):
        nh.cell(row=1, column=c).font = Font(bold=True)
    # 同一单位的几段挨在一处、按年份排好、编上序号 —— 抽取顺序堆着的话,
    # 一家单位的五个名字散在表里,谁先谁后全靠自己比对那串八位数字
    def _key(r):
        t = re.sub(r"\D", "", str(r.get("From", "")))
        return (str(r.get("Unit", "")), int(t.ljust(8, "0")[:8]) if t else 99999999)

    seq = {}
    for r in sorted(res["names"], key=_key):
        who = str(r.get("Unit", ""))
        seq[who] = seq.get(who, 0) + 1
        nh.append(["", seq[who], who, r.get("Name", ""), str(r.get("From", "")),
                   r.get("Remark", ""), r.get("Source", "")])
    nh.freeze_panes = "C2"
    for i, wid in enumerate([6, 5, 24, 26, 11, 34, 30], start=1):
        nh.column_dimensions[get_column_letter(i)].width = wid
    # 「序」是这张表最容易看岔的一格 —— 它不是行号,是这家单位的第几个名字。
    # 不写明白,谁也不知道那个 1、2、3 从哪儿来。
    nh_note = nh.cell(row=1, column=9,
                      value="↑ 一行 = 某单位某一段时间里叫什么。念法:"
                            "「自哪年起」那一年起,这家单位叫「当时名称」那个名字,到下一行那年为止。"
                            "「单位(今名)」是它如今的正名,一家单位的几行都写同一个 —— 那是钥匙,"
                            "不是它当时的名字。"
                            "「序」是这家单位的第几个名字,1 最早 —— 不是行号。"
                            "两家单位并成一家,不在这张表里,写进「待核」表的「创办」列。")
    nh_note.font = Font(italic=True, color="996600")

    # ---- 待核:核对用的那一张,原文摆在最后一列
    rv = wb.create_sheet("待核")
    rv_head = (["取否", "单位", "别名", "置信", "出处", "据以立论的原文", "行业", "产品",
                "始建", "终止", "创办", "地址", "区"] + stat_labels
               + ["统计年", "备注", "来路", "页"])
    rv.append(rv_head)
    for c in range(1, len(rv_head) + 1):
        rv.cell(row=1, column=c).font = Font(bold=True)
    for r in res["units"]:
        rv.append(["", r.get("Unit", ""), r.get("别名", ""),
                   r.get("confidence", ""), r.get("Source", ""),
                   r.get("evidence", ""), r.get("Industry", ""), r.get("Product", ""),
                   num(r.get("Start Date")), num(r.get("End Date")), r.get("Founder", ""),
                   r.get("Add.", ""), r.get("district", "")]
                  + [num(r.get(k)) for k in stat_keys]
                  + [r.get("统计年", ""), r.get("Remark", ""),
                     r.get("role", ""), r.get("page", "")])
    rv.freeze_panes = "D2"
    for i, wid in enumerate([6, 28, 24, 6, 30, 90, 10, 22, 11, 11, 30, 20, 7] + [9] * 8 + [8]
                            + [30, 6, 6], start=1):
        rv.column_dimensions[get_column_letter(i)].width = wid
    for row in rv.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].alignment = Alignment(wrap_text=False, vertical="top")

    widths = {sheet: [26, 10, 20, 11, 11, 40, 9, 22] + [9] * 8 + [40, 26],
              "Semi-Product": [6, 28, 20, 26, 24, 8, 11, 14, 30],
              "Comp-Product": [6, 30, 8, 12, 14, 30, 24, 26, 8, 20, 14, 16, 34],
              "Name-History": [6, 26, 30, 11, 40, 26]}
    for nm, ws_widths in widths.items():
        w = wb[nm]
        for i, wid in enumerate(ws_widths, start=1):
            w.column_dimensions[get_column_letter(i)].width = wid

    # 「待核」排在头一张,打开就停在它上头 —— 从前排在最末,一开文件停在
    # 「Fact and Comp-<城>」预览表上:那张 A 列是一列光秃秃的单位名,表头空着,
    # 又没有「取否」列。要核对的人第一眼看见的,恰恰是唯一改了不算数的那张。
    wb.move_sheet("待核", offset=-(len(wb.sheetnames) - 1))
    wb.active = wb.sheetnames.index("待核")
    for w in wb.worksheets:
        w.sheet_view.tabSelected = (w.title == "待核")

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    try:
        wb.save(path)
    except PermissionError:
        # Windows 上 Excel 开着文件就锁住它。抽了半天全在内存里,不能到末了
        # 一句 PermissionError 全丢了 —— 换个名字先存下来。
        stem, ext = os.path.splitext(path)
        alt, n = stem + "-新" + ext, 2
        while os.path.exists(alt):
            alt, n = "%s-新%d%s" % (stem, n, ext), n + 1
        wb.save(alt)
        log("！写不进 %s —— 多半正开在 Excel 里,文件被锁着。"
            % os.path.basename(path))
        log("  这一份改存到 %s;关掉 Excel 再跑一遍,才会写回原名。"
            % os.path.basename(alt))
        return alt
    log("Excel 已写到 %s" % path)
    log("  五张表:待核、Semi-Product、Comp-Product、Name-History、%s" % sheet)
    log("  打开就停在「待核」——「取否」在 A 列,写 y 的行才收。")
    log("  另外三张(Semi/Comp/Name-History)的「取否」也在 A 列,别漏了;")
    log("  「%s」是照原表体例排的预览,改它不算数。" % sheet)
    return path


# ---------------------------------------------------------------- 核过再读回来

KEEP_YES = {"y", "yes", "true", "1", "是", "要", "✓", "√"}

# 「待核」表头 → 抽取时用的字段名
REVIEW_COLS = {"取否": "keep", "来路": "role", "置信": "confidence", "页": "page",
               "序": None,
               "单位(今名)": "Unit", "当时名称": "Name", "自哪年起": "From",
               "别名": "别名",
               "单位": "Unit", "行业": "Industry", "产品": "Product",
               "始建": "Start Date", "终止": "End Date", "创办": "Founder",
               "地址": "Add.", "区": "district",
               "备注": "Remark", "出处": "Source", "统计年": "统计年",
               "据以立论的原文": "evidence"}
REVIEW_COLS.update({label: key for key, label in STAT_COLS})


DATE_COLS = ("Start Date", "End Date", "Time", "From")


def date_cell(v):
    """把手填的日期归到八位整数上。

    核对时是要直接改字的,不会有人记得「1958年建厂」该写成 19580000。
    四位年补成 19580000,六位补成 19580300,「1958年3月」照样认;Excel 把
    1958-03-01 存成了日期对象,也认。认不出就原样留着,不猜。"""
    if v is None or str(v).strip() == "":
        return ""
    if hasattr(v, "year") and hasattr(v, "month"):        # Excel 的日期格
        return "%04d%02d%02d" % (v.year, v.month, getattr(v, "day", 0) or 0)
    t = str(v).strip()
    if re.fullmatch(r"\d{8}", t):
        return t
    if re.fullmatch(r"\d{6}", t):
        return t + "00"
    if re.fullmatch(r"\d{4}", t) and 1800 <= int(t) <= 2100:
        return t + "0000"
    return cndate.parse(t) or t


def _yes(v):
    return str("" if v is None else v).strip().lower() in KEEP_YES


def _sheet_rows(ws):
    """一张表读成 [{表头: 值}],整行空的跳过。"""
    head = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        r = {}
        for h, v in zip(head, row):
            if not h:
                continue
            # 表头行末尾那句以「↑」起头的话是写给人看的注,不是一列数据
            if str(h).startswith("↑"):
                continue
            k = REVIEW_COLS.get(h, h)
            if k is None:      # 「序」是给人看的次序,不是字段
                continue
            r[k] = date_cell(v) if k in DATE_COLS else ("" if v is None else v)
        out.append(r)
    return out


def merge_by_name(rows):
    """核过之后名字写成一样的,就当同一家,合成一行。

    这正是核名字要干的事:「四机部15所」「电子部15所」「电子部第15所」本是
    一个所,随部委改制换了牌子。在表里把它们都改成一个名字,这里就并起来,
    各自的出处一并留着 —— 并了以后仍要查得回去是哪一节说的。"""
    out, idx, merged = [], {}, 0
    for r in rows:
        nm = str(r.get("Unit", "")).strip()
        if not nm:
            continue
        if nm not in idx:
            idx[nm] = dict(r, Unit=nm)
            out.append(idx[nm])
            continue
        merged += 1
        base = idx[nm]
        for k, v in r.items():
            if v not in ("", None) and base.get(k) in ("", None):
                base[k] = v
        for k in ("Source", "Remark"):
            a, b = str(base.get(k, "") or ""), str(r.get(k, "") or "")
            if b and b not in a:
                base[k] = (a + "；" + b) if a else b
        # 并进来的那些名字要留着 —— 挑一个当正名,不等于别的就不算数了
        alias = [x.strip() for x in
                 (str(base.get("别名", "") or "") + "、" + str(r.get("别名", "") or "")).split("、")]
        base["别名"] = "、".join(dict.fromkeys(
            x for x in alias if x and x != nm))
    return out, merged


def read_review(path):
    """把核过的工作簿读回来:四张表里「取否」写了 y 的行。

    落笔的地方只有一处 —— Excel。keep 在那儿打,认错的字也在那儿改,读回来
    的就是你改过的样子。TSV 只当留底,不再回头去读:两处都能改,改了哪一处
    算数就说不清了。

    返回 (四张表的行, 城市)。城市从「Fact and Comp-北京」这类表名上取。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    city = ""
    for name in wb.sheetnames:
        if name.startswith("Fact and Comp-"):
            city = name[len("Fact and Comp-"):]
            break

    bundle, seen = {}, {}
    for tag, sheet in (("units", "待核"), ("semi", "Semi-Product"),
                       ("comp", "Comp-Product"), ("names", "Name-History")):
        if sheet not in wb.sheetnames:
            bundle[tag], seen[tag] = [], 0
            continue
        rows = _sheet_rows(wb[sheet])
        seen[tag] = len(rows)
        kept = []
        for r in rows:
            if not _yes(r.pop("keep", "")):
                continue
            r.pop("evidence", None)
            if tag == "units" and city and not r.get("City"):
                r["City"] = city
            if tag == "names" and r.get("From") != "":
                r["From"] = str(r["From"])
            kept.append(r)
        if tag == "units":
            kept, seen["merged"] = merge_by_name(kept)
        bundle[tag] = kept
    return bundle, city, seen
