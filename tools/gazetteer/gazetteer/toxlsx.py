# -*- coding: utf-8 -*-
"""过了人眼的记录 → 追加进 CN_Electronic_Industry.xlsx。

**只追加,不改动既有行。** 动手之前先把原文件另存一份带时间戳的备份。
表头按名字认列,不按字母位置认 —— 将来在表尾添 `Name EN` / `Lat` / `Lng`
也不会串行(见 src/xlsxio.js 的读法)。
"""

import io
import os
import re
import shutil

from openpyxl.utils import get_column_letter
from datetime import datetime

SHEET_UNITS = "Fact and Comp-Shanghai"
SHEET_SEMI = "Semi-Product"
SHEET_COMP = "Comp-Product"
SHEET_NAMES = "Name-History"

STAT_LABELS = [("staff", "职工总数"), ("tech", "技术人员"), ("plant", "厂房面积"),
               ("floor", "建筑面积"), ("assets", "固定资产"), ("output", "工业总产值"),
               ("sales", "销售收入"), ("profit", "实现利润")]


def _open(path):
    import openpyxl
    return openpyxl.load_workbook(path)


def _headers(ws, rows=2):
    """{表头文字: 列号},两行表头一并收。"""
    idx = {}
    for r in range(1, rows + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            key = str(v).strip()
            if key and key not in idx:
                idx[key] = c
    return idx


def _num(v):
    """数字照数字写,日期写成八位整数,其余照录字符串。"""
    s = str(v if v is not None else "").strip()
    if s == "":
        return None
    if re.fullmatch(r"\d{8}", s):
        return int(s)
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return s


ALIAS_SPLIT = re.compile(r"[、,，;；/／|]")


def _bare(v):
    """判重用的样子:去掉括号与空白。「上海无线电十九厂(上无十九)」→ 上海无线电十九厂"""
    return re.sub(r"\s+", "", re.sub(r"[（(][^）)]*[）)]", "", str(v or ""))).strip()


def _names_of(name_cell, alias_cell=""):
    """这一行占着哪些名字 —— 正名、名字里括号中的、以及「别名」列里的。"""
    out = {_bare(name_cell)} if _bare(name_cell) else set()
    for grp in re.findall(r"[（(]([^）)]*)[）)]", str(name_cell or "")):
        out |= {x.strip() for x in ALIAS_SPLIT.split(grp) if x.strip()}
    out |= {x.strip() for x in ALIAS_SPLIT.split(str(alias_cell or "")) if x.strip()}
    return out


def _existing_names(ws, first_data_row, alias_col=None):
    """表里已经占着的名字。别名也算 —— 「四机部15所」与「华北计算技术研究所」
    是一家,凭正名比对,同一家会当成两家收两次。"""
    out = set()
    for r in range(first_data_row, ws.max_row + 1):
        alias = ws.cell(row=r, column=alias_col).value if alias_col else ""
        out |= _names_of(ws.cell(row=r, column=1).value, alias)
    return out


def _existing_keys(ws, cols, h, first_data_row=2):
    """器件、整机、沿革三张表的判重钥匙 —— 原先一道也没有,跑两遍就多一份。"""
    seen = set()
    for r in range(first_data_row, ws.max_row + 1):
        key = tuple(_bare(ws.cell(row=r, column=h[c]).value) if c in h else ""
                    for c in cols)
        if any(key):
            seen.add(key)
    return seen


def _last_row(ws, col=1, start=1):
    last = start - 1
    for r in range(start, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "")
               for c in range(1, ws.max_column + 1)):
            last = r
    return last


def append(xlsx_path, units=(), semi=(), comp=(), names=(), backup=True,
           allow_dup=False, log=print):
    wb = _open(xlsx_path)
    report = {"backup": "", "units": 0, "semi": 0, "comp": 0, "names": 0,
              "skipped": [], "near": []}

    if backup:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        bak = "%s.backup-%s.xlsx" % (os.path.splitext(xlsx_path)[0], stamp)
        shutil.copy2(xlsx_path, bak)
        report["backup"] = bak
        log("原表已备份到 %s" % os.path.basename(bak))

    # ---- 厂所名录(两行表头,正文自第 3 行起)
    if units:
        ws = wb[SHEET_UNITS]
        h = _headers(ws, 2)
        have = _existing_names(ws, 3, alias_col=h.get("别名"))
        row = _last_row(ws, start=3) + 1
        for r in units:
            nm = str(r.get("Unit") or r.get("name") or "").strip()
            if not nm:
                continue
            mine = _names_of(nm, r.get("别名", ""))
            if not allow_dup and (mine & have):
                report["skipped"].append(
                    "%s(表内已有「%s」)" % (nm, sorted(mine & have)[0])
                    if _bare(nm) not in have else nm)
                continue
            ws.cell(row=row, column=1, value=nm)
            if str(r.get("别名") or "").strip():
                _ensure_column(ws, "别名", header_row=1)
                h = _headers(ws, 2)
            for label in ("Industry", "Product", "Start Date", "End Date", "Founder",
                          "City", "Add.", "Remark", "Source", "Name EN", "Lat", "Lng",
                          "别名"):
                if label in h and r.get(label) not in (None, ""):
                    ws.cell(row=row, column=h[label], value=_num(r[label]))
            for key, label in STAT_LABELS:
                if label in h and r.get(key) not in (None, ""):
                    ws.cell(row=row, column=h[label], value=_num(r[key]))
            # 区名一直是抽出来了却没写下 —— 于是一市之内所有的点都落在市中心,
            # 叠成一坨。有了区,至少能落到区一级(见 src/xlsxio.js 的 district 兜底)
            if r.get("district") not in (None, ""):
                _ensure_column(ws, "区", header_row=1)
                h = _headers(ws, 2)
                ws.cell(row=row, column=h["区"]).value = str(r["district"]).strip()
            # 数字是哪一年的 —— 志书各章截取的年份不一致(上海多是 1990,
            # 北京第四篇是 1995),不记下来,一个数字就等于没说
            if r.get("统计年") not in (None, ""):
                _ensure_column(ws, "统计年", header_row=1)
                h = _headers(ws, 2)
                ws.cell(row=row, column=h["统计年"], value=_num(r["统计年"]))
            have |= mine
            row += 1
            report["units"] += 1

    def _append_flat(sheet, cols, rows, tag, text_cols=(), ensure=(), dedup_on=()):
        if not rows:
            return
        ws = wb[sheet]
        for label in ensure:
            if any(r.get(label) not in (None, "") for r in rows):
                _ensure_column(ws, label)
        h = _headers(ws, 1)
        keys = set() if allow_dup else _existing_keys(ws, dedup_on, h)
        row = _last_row(ws, start=2) + 1
        for r in rows:
            k = tuple(_bare(r.get(c)) for c in dedup_on)
            if any(k) and k in keys:
                report["skipped"].append("%s:%s" % (tag, k[0]))
                continue
            if any(k):
                keys.add(k)
            wrote = False
            for label in cols:
                if label in h and r.get(label) not in (None, ""):
                    v = str(r[label]) if label in text_cols else _num(r[label])
                    ws.cell(row=row, column=h[label], value=v)
                    wrote = True
            if wrote:
                row += 1
                report[tag] += 1

    _append_flat(SHEET_SEMI, ["Product", "别名", "Research Insti", "Factory", "产量", "Time",
                             "Personnel", "Remark"],
                 semi, "semi", ensure=("产量", "别名", "Research Insti"),
                 dedup_on=("Product", "Factory", "Time"))
    # 「用户」是原表没有的一列 —— 机器交到谁手里用,记在这儿(见 src/xlsxio.js)
    if comp:
        from .extract import model_key
        ws = wb[SHEET_COMP]
        hc = _headers(ws, 1)
        if "Product" in hc:
            core = {}
            for rr in range(2, ws.max_row + 1):
                pv = ws.cell(row=rr, column=hc["Product"]).value
                k = model_key(pv) if pv else ""
                if k:
                    core.setdefault(k, set()).add(str(pv).strip())
            for r in comp:
                pv = str(r.get("Product") or "").strip()
                k = model_key(pv)
                old_names = core.get(k, set()) - {pv} if k else set()
                if old_names:
                    report["near"].append((pv, "、".join(sorted(old_names))))

    _append_flat(SHEET_COMP, ["Product", "字长", "内存", "Speed（次秒）", "Research Insti",
                              "Factory", "用户", "产量", "别名", "Time", "Personnel", "Remark"], comp, "comp",
                 ensure=("用户", "产量", "别名"), dedup_on=("Product", "Time"))
    # Name-History 的 From 一列,原表存的是文本(见 src/xlsxio.js 的 exportWorkbook),照旧
    _append_flat(SHEET_NAMES, ["Unit", "Name", "From", "Name EN", "Remark", "Source"],
                 names, "names", text_cols=("From",), dedup_on=("Unit", "Name", "From"))
    _tidy_after = report["names"] > 0

    try:
        wb.save(xlsx_path)
        # 追加是往表尾堆,新来的几段跟同一单位原有的隔着几十行 —— 存完就地理一遍
        if _tidy_after:
            tidy_names(xlsx_path)
    except PermissionError:
        # 这一份是总表,不能改名另存 —— 换了名字就不是那本工作簿了
        raise SystemExit(
            "写不进 %s —— 多半正开在 Excel 里,文件被锁着。\n"
            "关掉 Excel 再跑一遍。原表还是原样,备份也还在:%s"
            % (os.path.basename(xlsx_path), report["backup"] or "(这次没备份)"))
    return report


def read_known(xlsx_path):
    """把表里已有的单位名读出来(连括号里的别名),供抽取时辨认旧识。"""
    if not os.path.exists(xlsx_path):
        return {}
    wb = _open(xlsx_path)
    ws = wb[SHEET_UNITS]
    known = {}
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        raw = str(v).strip()
        canon = re.sub(r"[（(][^）)]*[）)]", "", raw).strip()
        m = re.search(r"[（(]([^）)]*)[）)]", raw)
        known[canon] = [raw] + ([m.group(1).strip()] if m else [])
    return known


def read_aliases(geocode_js):
    """把 src/geocode.js 的 ALIASES 读出来。

    「上无十三」「上五十三」这类简称不合单位名的体例,正则认不出来,
    只能靠这张人工对照表。它已经在仓库里,没有理由再抄一份。"""
    if not os.path.exists(geocode_js):
        return {}
    with open(geocode_js, encoding="utf-8") as f:
        src = f.read()
    m = re.search(r"export\s+const\s+ALIASES\s*=\s*\{(.*?)\n\};", src, re.S)
    if not m:
        return {}
    out = {}
    for line in m.group(1).splitlines():
        mm = re.match(r'\s*"([^"]+)"\s*:\s*\[([^\]]*)\]', line)
        if mm:
            out[mm.group(1)] = re.findall(r'"([^"]+)"', mm.group(2))
    return out


def merge_known(xlsx_path, geocode_js=None):
    known = read_known(xlsx_path)
    for canon, aliases in (read_aliases(geocode_js) if geocode_js else {}).items():
        known.setdefault(canon, [])
        for a in aliases:
            if a not in known[canon]:
                known[canon].append(a)
    return known


def read_places(geocode_js):
    """src/geocode.js 的 PLACES 里已经有落点的单位名。"""
    if not os.path.exists(geocode_js):
        return set()
    with open(geocode_js, encoding="utf-8") as f:
        src = f.read()
    m = re.search(r"export\s+const\s+PLACES\s*=\s*\{(.*?)\n\};", src, re.S)
    return set(re.findall(r'^\s*"([^"]+)"\s*:', m.group(1), re.M)) if m else set()


# ============================================================
#  以下供 vault.py 用:整本读出、按行改回
#  ------------------------------------------------------------
#  append() 只管往表尾添行;库里改过的字段要写回**原来那一行**,
#  故另备一套按表头认列、按行号定位的读写。
# ============================================================

UNIT_LABELS = ["Industry", "Product", "Start Date", "End Date", "Founder", "City", "Add.",
               "Remark", "Source", "Name EN", "Lat", "Lng"]


def read_units_full(xlsx_path):
    """厂所名录整本读出,每行带着它在表里的行号,回写时据以定位。"""
    wb = _open(xlsx_path)
    ws = wb[SHEET_UNITS]
    h = _headers(ws, 2)
    out = []
    for r in range(3, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None or not str(raw).strip():
            continue
        rec = {"_row": r, "raw": str(raw).strip()}
        for label in UNIT_LABELS:
            rec[label] = ws.cell(row=r, column=h[label]).value if label in h else None
        for key, label in STAT_LABELS:
            rec[key] = ws.cell(row=r, column=h[label]).value if label in h else None
        rec["统计年"] = ws.cell(row=r, column=h["统计年"]).value if "统计年" in h else None
        rec["district"] = ws.cell(row=r, column=h["区"]).value if "区" in h else None
        out.append(rec)
    return out


def read_flat(xlsx_path, sheet, cols):
    wb = _open(xlsx_path)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    h = _headers(ws, 1)
    out = []
    for r in range(2, ws.max_row + 1):
        rec = {"_row": r}
        got = False
        for label in cols:
            v = ws.cell(row=r, column=h[label]).value if label in h else None
            rec[label] = v
            if v not in (None, ""):
                got = True
        if got:
            out.append(rec)
    return out


def read_name_history(xlsx_path):
    return read_flat(xlsx_path, SHEET_NAMES, ["Unit", "Name", "From", "Name EN", "Remark", "Source"])


def read_semi(xlsx_path):
    return read_flat(xlsx_path, SHEET_SEMI, ["Product", "Factory", "Time", "Personnel", "Remark"])


def read_comp(xlsx_path):
    return read_flat(xlsx_path, SHEET_COMP, ["Product", "字长", "内存", "Speed（次秒）",
                                             "Research Insti", "Factory", "Time",
                                             "Personnel", "Remark"])


def _ensure_column(ws, label, header_row=1):
    """表尾按需添一列(Name EN / Lat / Lng 原表没有,库里填了才添)。"""
    h = _headers(ws, 2)
    if label in h:
        return h[label]
    col = ws.max_column + 1
    ws.cell(row=header_row, column=col, value=label)
    return col


def update_units(xlsx_path, changes, backup=True, log=print):
    """把库里改动的字段写回原行。

    changes: [{"_row": 3, "名称": "...", "fields": {"Industry": "半导体", ...}}]
    只动列出来的格子,别的一律不碰。"""
    wb = _open(xlsx_path)
    ws = wb[SHEET_UNITS]
    bak = ""
    if backup:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        bak = "%s.backup-%s.xlsx" % (os.path.splitext(xlsx_path)[0], stamp)
        shutil.copy2(xlsx_path, bak)
        log("原表已备份到 %s" % os.path.basename(bak))

    n = 0
    for ch in changes:
        row = ch["_row"]
        if "名称" in ch:
            ws.cell(row=row, column=1, value=ch["名称"])
            n += 1
        for label, val in (ch.get("fields") or {}).items():
            col = _ensure_column(ws, label)
            ws.cell(row=row, column=col, value=(None if val in ("", None) else _num(val)))
            n += 1
    wb.save(xlsx_path)
    return {"cells": n, "backup": bak}


def rewrite_name_history(xlsx_path, rows, backup=False, log=print):
    """名称沿革整表重写 —— 库里改的是「某单位的全部段落」,逐格补丁反而易错。"""
    wb = _open(xlsx_path)
    if SHEET_NAMES not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAMES)
        ws.append(["Unit", "Name", "From", "Name EN", "Remark", "Source"])
    ws = wb[SHEET_NAMES]
    # 原表没有 Name EN 一列;库里填了英文名就把这一列添上,不能悄悄丢掉
    if any(str(r.get("Name EN") or "").strip() for r in rows):
        _ensure_column(ws, "Name EN")
    h = _headers(ws, 1)
    order = sorted(h, key=lambda k: h[k])
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, r in enumerate(rows, start=2):
        for label in order:
            if label not in h:
                continue
            v = r.get(label, "")
            ws.cell(row=i, column=h[label],
                    value=(None if v in ("", None) else (str(v) if label == "From" else _num(v))))
    wb.save(xlsx_path)
    return len(rows)


# ---------------------------------------------------------------- 比两份工作簿

# 每张表拿什么当钥匙认「同一行」
DIFF_KEYS = {
    SHEET_UNITS: (("Unit",), 3, 2),
    SHEET_COMP: (("Product", "Time"), 2, 1),
    SHEET_SEMI: (("Product", "Factory", "Time"), 2, 1),
    SHEET_NAMES: (("Unit", "Name", "From"), 2, 1),
}


def _sheet_rows(wb, sheet, key_cols, first_row, header_rows):
    """{钥匙: {表头: 值}} —— 认不出钥匙的那些行另存一份,免得默默丢掉。"""
    if sheet not in wb.sheetnames:
        return {}, []
    ws = wb[sheet]
    h = _headers(ws, header_rows)
    name_col = 1 if sheet == SHEET_UNITS else None
    out, odd = {}, []
    for r in range(first_row, ws.max_row + 1):
        rec = {}
        for label, c in h.items():
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                rec[label] = str(v).strip()
        if name_col:
            v = ws.cell(row=r, column=name_col).value
            if v not in (None, ""):
                rec["Unit"] = str(v).strip()
        if not rec:
            continue
        key = tuple(_bare(rec.get(c, "")) for c in key_cols)
        if not any(key):
            odd.append(rec)
            continue
        out.setdefault(key, rec)
    return out, odd


def diff_workbooks(old_path, new_path, skip_cols=("序", "至")):
    """两份工作簿差在哪儿 —— `.xlsx` 是二进制,git 只会说「变了」,不说变了什么。

    按钥匙认同一行(厂所认名字,产品认型号+年份),逐格比。「序」「至」是
    算出来的,不算数据变动,默认略过。"""
    import openpyxl
    a = openpyxl.load_workbook(old_path, data_only=True)
    b = openpyxl.load_workbook(new_path, data_only=True)
    out = {}
    for sheet, (key_cols, first_row, header_rows) in DIFF_KEYS.items():
        ao, a_odd = _sheet_rows(a, sheet, key_cols, first_row, header_rows)
        bo, b_odd = _sheet_rows(b, sheet, key_cols, first_row, header_rows)
        added = [bo[k] for k in bo if k not in ao]
        gone = [ao[k] for k in ao if k not in bo]
        changed = []
        for k in ao:
            if k not in bo:
                continue
            fields = []
            for label in sorted(set(ao[k]) | set(bo[k])):
                if label in skip_cols:
                    continue
                x, y = ao[k].get(label, ""), bo[k].get(label, "")
                if x != y:
                    fields.append((label, x, y))
            if fields:
                changed.append(("·".join(x for x in k if x), fields))
        if added or gone or changed:
            out[sheet] = {"added": added, "gone": gone, "changed": changed,
                          "odd": len(a_odd) + len(b_odd)}
    return out


# ---------------------------------------------------------------- 理沿革表

def _ymd(v):
    """八位日期 → 可比的整数;认不出的排在最后。"""
    t = re.sub(r"\D", "", str(v or ""))
    return int(t.ljust(8, "0")[:8]) if t else 99999999


def _fmt_ym(v):
    """19501100 → 1950.11;只知道年的写 1950。"""
    t = re.sub(r"\D", "", str(v or ""))
    if len(t) != 8:
        return str(v or "")
    y, m, d = t[:4], t[4:6], t[6:]
    if m == "00":
        return y
    return "%s.%s" % (y, m) if d == "00" else "%s.%s.%s" % (y, m, d)


def tidy_names(xlsx_path, dry_run=False):
    """把沿革表理一理:同一单位的几段挨在一处,按年份先后排好,重编「序」,
    并按下一段的启用年补上「至」。

    这张表一行是一个**名段**:某单位从某年起叫什么。可原先只有 Unit/Name/From
    三列,行又是按抽取顺序堆着的 —— 一家单位的五个名字散在表里,谁先谁后
    全靠自己比对那串八位数字。「序」与「至」都是算出来的,手工插过行就会
    对不上,所以单拎出这一道,随时可以再理一遍。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAMES not in wb.sheetnames:
        return {"moved": 0, "rows": 0}
    ws = wb[SHEET_NAMES]
    h = _headers(ws, 1)
    if "Unit" not in h or "From" not in h:
        return {"moved": 0, "rows": 0}

    ncol = ws.max_column
    body = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
        if any(v not in (None, "") for v in vals):
            body.append(vals)
    before = [tuple(x) for x in body]

    ui, fi = h["Unit"] - 1, h["From"] - 1
    body.sort(key=lambda v: (str(v[ui] or ""), _ymd(v[fi])))

    # 读起来顺的次序:序 | 单位 | 名称 | 起 | 至 | 关系 | 其余照旧
    head = [ws.cell(row=1, column=c).value for c in range(1, ncol + 1)]
    lead = ["序", "Unit", "Name", "From", "至", "关系"]
    order = [head.index(x) if x in head else None for x in lead]
    order += [i for i, v in enumerate(head) if v not in lead]
    rows = []
    for v in body:
        rows.append([None if i is None else (v[i] if i < len(v) else None) for i in order])
    head = [lead[i] if i < len(lead) else head[order[i]] for i in range(len(order))]

    si, ui2, fi2, ti = 0, 1, 3, 4
    for i, v in enumerate(rows):
        same = [j for j, w in enumerate(rows) if str(w[ui2] or "") == str(v[ui2] or "")]
        v[si] = same.index(i) + 1
        nxt = [j for j in same if j > i]
        # 「至」是下一段启用那一年 —— 名字换在那一天,前一个名字用到那一天为止
        v[ti] = rows[nxt[0]][fi2] if nxt else None

    # 一律写 .value,不走 cell(..., value=x):openpyxl 里 value=None 当作
    # 「没给值」,格子里的旧字原样留着 —— 该空的那一格于是留着上一轮的内容
    for c, label in enumerate(head, start=1):
        ws.cell(row=1, column=c).value = label
    for i, v in enumerate(rows):
        for c in range(1, len(head) + 1):
            ws.cell(row=i + 2, column=c).value = v[c - 1]
    for r in range(len(rows) + 2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None
    for c, wid in zip(range(1, len(head) + 1), (5, 24, 26, 10, 10, 8, 34, 30)):
        ws.column_dimensions[get_column_letter(c)].width = wid
    ws.freeze_panes = "C2"

    moved = sum(1 for a, b in zip(before, [tuple(x) for x in body]) if a != b) if before else 0
    if not dry_run:
        wb.save(xlsx_path)
    return {"moved": moved, "rows": len(rows)}


# ---------------------------------------------------------------- 找重复

def _sheet_dups(ws, key_cols, label_col, header_row=1, first_data_row=2):
    """按钥匙分组,返回 {钥匙: [(行号, 标签)]},只留下不止一行的。"""
    h = _headers(ws, header_row)
    if not all(c in h for c in key_cols):
        return {}
    groups = {}
    for r in range(first_data_row, ws.max_row + 1):
        key = tuple(_bare(ws.cell(row=r, column=h[c]).value) for c in key_cols)
        if not any(key):
            continue
        label = ws.cell(row=r, column=h.get(label_col, 1)).value
        groups.setdefault(key, []).append((r, str(label or "")))
    return {k: v for k, v in groups.items() if len(v) > 1}


def report_dups(xlsx_path):
    """把工作簿里疑似重复的地方找出来 —— 只报,一个格子也不动。

    分两等:**一模一样**的(整条钥匙相同)多半是追加了两遍,该删;
    **像是一回事**的(单位名字有交叠、型号内核相同)得你自己看 ——
    「DJS-130」与「DJS-130B」型号内核一样,却是两台机器。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {"exact": [], "similar": []}

    ws = wb[SHEET_UNITS]
    h = _headers(ws, 2)
    seen = {}
    for r in range(3, ws.max_row + 1):
        nm = ws.cell(row=r, column=1).value
        if not nm:
            continue
        alias = ws.cell(row=r, column=h["别名"]).value if "别名" in h else ""
        for one in _names_of(nm, alias):
            seen.setdefault(one, []).append((r, str(nm)))
    for one, rows in seen.items():
        if len(rows) > 1:
            who = sorted({x[1] for x in rows})
            kind = "exact" if len(who) == 1 else "similar"
            out[kind].append(("厂所", one, [x[0] for x in rows], "、".join(who)))

    # 名字差一截的两家 ——「北大方正集团公司」与「北京北大方正集团公司」、
    # 「联想计算机集团公司」与「北京联想计算机集团公司」。判重只比名字是否相同,
    # 这种一个是另一个的一截,两边都过关,于是同一家在名录里坐了两行。
    # 多出来的这一截若是「分厂」「分公司」一类,那本就是另一家,不算重复
    BRANCH = re.compile(r"(分厂|分公司|分所|附属工厂|实习厂|车间|工场)$")
    short = sorted({(one, who) for one, rows in seen.items() for _r, who in rows},
                   key=lambda x: len(x[0]))
    for i, (a, wa) in enumerate(short):
        if len(a) < 5:
            continue
        for b, wb_ in short[i + 1:]:
            if a == b or not (b.endswith(a) or b.startswith(a)):
                continue
            if len(b) - len(a) > 4 or BRANCH.search(b):
                continue
            ra = {x[0] for x in seen.get(a, [])}
            rb = {x[0] for x in seen.get(b, [])}
            if ra == rb:          # 同一行的正名与别名,不是两家
                continue
            out["similar"].append(("厂所", "%s ⊂ %s" % (a, b), sorted(ra | rb),
                                   "%s ｜ %s —— 名字差一截,是不是同一家?" % (wa, wb_)))

    for sheet, key_cols, label in (
            (SHEET_COMP, ("Product", "Time"), "Product"),
            (SHEET_SEMI, ("Product", "Factory", "Time"), "Product"),
            (SHEET_NAMES, ("Unit", "Name", "From"), "Unit")):
        if sheet not in wb.sheetnames:
            continue
        for key, rows in _sheet_dups(wb[sheet], key_cols, label).items():
            out["exact"].append((sheet, "·".join(x for x in key if x),
                                 [x[0] for x in rows], rows[0][1]))

    # 型号内核相同、写法不同的整机 —— 只提醒,不当重复
    if SHEET_COMP in wb.sheetnames:
        from .extract import model_key
        ws = wb[SHEET_COMP]
        h = _headers(ws, 1)
        if "Product" in h:
            by = {}
            for r in range(2, ws.max_row + 1):
                p = ws.cell(row=r, column=h["Product"]).value
                k = model_key(p) if p else ""
                if k:
                    by.setdefault(k, []).append((r, str(p).strip()))
            for k, rows in by.items():
                names = sorted({x[1] for x in rows})
                if len(names) > 1:
                    detail = []
                    for rr, nm in rows:
                        t = ws.cell(row=rr, column=h["Time"]).value if "Time" in h else ""
                        f = (ws.cell(row=rr, column=h["Factory"]).value
                             if "Factory" in h else "") or ""
                        detail.append("%s(%s%s)" % (nm, t or "年份未详",
                                                    ("·" + str(f)) if f else ""))
                    out["similar"].append((SHEET_COMP, k, [x[0] for x in rows],
                                           " ｜ ".join(detail)))
    return out


# ---------------------------------------------------------------- 手改之后验一验

# 「1966年（后改名…」被从中切开,「6年」当成了纪元六年 —— 折出 1906。
# 中国电子工业没有 1900 年代的事,这十年里的年份一概是这么来的。
SLICED_YEAR = range(1900, 1910)

# 产品表没有出处列,书名混在备注末尾的那截出处里:「…北京工业志·电子志·第三章…」
BOOK_IN_TEXT = re.compile(r"([一-鿿]{2,12}志)[·・]")


def _book_of(*cells):
    """这一行是哪本志抄来的 —— 判断「市计算机技术研究所」是哪个市的,全靠它。"""
    for v in cells:
        m = BOOK_IN_TEXT.search(str(v or ""))
        if m:
            return m.group(1)
    return ""


def _year_trouble(v, label):
    """日期格子有没有毛病 —— 有就说一句,没有就返回 None。"""
    t = re.sub(r"\s+", "", str(v))
    if not re.fullmatch(r"\d{8}", t):
        return "%s 不是八位日期:%r —— 只知道年份写 19580000" % (label, v)
    y = int(t[:4])
    if not (1800 <= y <= 2100):
        return "%s 的年份不像话:%d" % (label, y)
    if y in SLICED_YEAR:
        return ("%s = %d,多半是「1966年（…」被从中切开算出来的,"
                "回原文核一核" % (label, y))
    return None


def accepted_path(xlsx_path):
    """《已核》摆在工作簿旁边,同名 —— 两个一起提交,记录才跟着数据走。"""
    return os.path.splitext(xlsx_path)[0] + ".已核.tsv"


def load_accepted(xlsx_path):
    """从前看过、认下了的那些毛病 —— 再报一遍只会把新的埋掉。"""
    path = accepted_path(xlsx_path)
    if not os.path.exists(path):
        return set()
    out = set()
    with io.open(path, encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue
            col = line.rstrip("\n").split("\t")
            if len(col) >= 2:
                out.add((col[0], col[1]))
    return out


def save_accepted(xlsx_path, findings):
    """把眼下报的这些一律记作「看过了」。

    写的是当下的全部,所以已经改好、不再报的那些会自己掉出去 ——
    《已核》不会越积越长,始终只是「此刻认下的这些」。"""
    path = accepted_path(xlsx_path)
    rows = sorted({(k, key, why) for k, _w, why, key in findings})
    with io.open(path, "w", encoding="utf-8", newline="") as f:
        f.write("# 看过了、认下了的毛病 —— gaz verify 不再报。\n")
        f.write("# 要全部重看一遍:gaz verify --all\n")
        f.write("# 改好一条,它自己就不报了;这份名单在下次 --accept 时重写。\n")
        f.write("# 哪一类\t记号\t当时报的是什么\n")
        for kind, key, why in rows:
            f.write("%s\t%s\t%s\n" % (kind, key, why.replace("\t", " ")))
    return path, len(rows)


def verify(xlsx_path):
    """手改过工作簿之后,看看有没有改坏。只报,一个格子也不动。

    Excel 改起来顺手,坏起来也无声无息:日期写成 1958、坐标只填了一半、
    整机的研制单位打错一个字 —— 站点照读不误,只是那条连线从此连不上,
    而表面上什么也看不出来。

    返回 [(哪一类, 哪一行, 怎么了, 记号)]。分类是为了让「出处空着」这种
    成片的旧账,别把「研制单位打错字」这种一处一处的伤埋掉。「记号」是这条
    毛病的身份,不含行号 —— 行会挪,毛病还是那个毛病,拿它跟《已核》比对。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    bad = []

    ws = wb[SHEET_UNITS]
    h = _headers(ws, 2)
    names = set()
    for r in range(3, ws.max_row + 1):
        nm = ws.cell(row=r, column=1).value
        if not nm:
            continue
        nm = str(nm).strip()
        names |= _names_of(nm, ws.cell(row=r, column=h["别名"]).value if "别名" in h else "")
        where = "%s 第%d行" % (nm, r)
        key = nm

        for label in ("Start Date", "End Date"):
            if label not in h:
                continue
            v = ws.cell(row=r, column=h[label]).value
            if v in (None, ""):
                continue
            why = _year_trouble(v, label)
            if why:
                bad.append(("日期", where, why, "%s·%s" % (key, label)))

        lat = ws.cell(row=r, column=h["Lat"]).value if "Lat" in h else None
        lng = ws.cell(row=r, column=h["Lng"]).value if "Lng" in h else None
        if (lat in (None, "")) != (lng in (None, "")):
            bad.append(("坐标", where, "经纬度只填了一半 —— 两个都填,或者都空着", key))
        else:
            try:
                if lat not in (None, "") and not (3 <= float(lat) <= 54 and 73 <= float(lng) <= 136):
                    bad.append(("坐标", where,
                                "经纬度不在中国范围内:%s, %s(是不是填反了?)" % (lat, lng), key))
            except (TypeError, ValueError):
                bad.append(("坐标", where, "经纬度不是数字:%r, %r" % (lat, lng), key))

        if "Source" in h and not str(ws.cell(row=r, column=h["Source"]).value or "").strip():
            bad.append(("出处", where, "没写出处", key))

    # 沿革表:年份不像话的、以及孤零零一条「自己改名叫自己」的
    if SHEET_NAMES in wb.sheetnames:
        w = wb[SHEET_NAMES]
        hh = _headers(w, 1)
        rows, per = [], {}
        for r in range(2, w.max_row + 1):
            u = _bare(w.cell(row=r, column=hh["Unit"]).value) if "Unit" in hh else ""
            n = _bare(w.cell(row=r, column=hh["Name"]).value) if "Name" in hh else ""
            f = w.cell(row=r, column=hh["From"]).value if "From" in hh else None
            if not u:
                continue
            rows.append((r, u, n, f))
            per[u] = per.get(u, 0) + 1
        for r, u, n, f in rows:
            where = "%s 第%d行 %s" % (SHEET_NAMES, r, u)
            if f in (None, ""):
                # 一段名字没有启用年,就不知道它管哪一截 —— 站点按年份挑名字,
                # 挑不着它;这一行等于没写。前身没丢,它在「创办」列里。
                bad.append(("沿革", where, "没有启用年 —— 一段名字系不上年份,"
                            "这一行立不住(前身仍在「创办」列里)", key))
            else:
                why = _year_trouble(f, "启用年")
                if why:
                    bad.append(("日期", where, why, key))
            # 「甲厂 → 甲厂」是改名链的末一段,前头得有段别的名字才立得住
            if n == u and per.get(u, 0) < 2:
                bad.append(("沿革", where, "只此一条,却是「改名叫自己」—— 不载信息", key))

        # 「序」「至」是算出来的:手工插过行、改过年份,就跟实际对不上了
        if "序" in hh and rows:
            fresh = sorted(rows, key=lambda t: (t[1], _ymd(t[3])))
            stale = 0
            for i, (r, u, n, f) in enumerate(fresh):
                same = [x for x in fresh if x[1] == u]
                want_no = same.index((r, u, n, f)) + 1
                pos = same.index((r, u, n, f))
                want_to = same[pos + 1][3] if pos + 1 < len(same) else None
                got_no = w.cell(row=r, column=hh["序"]).value
                got_to = w.cell(row=r, column=hh["至"]).value if "至" in hh else None
                if (str(got_no or "") != str(want_no)
                        or _ymd(got_to) != _ymd(want_to)):
                    stale += 1
            if stale:
                bad.append(("沿革", "%s 全表" % SHEET_NAMES,
                            "「序」或「至」有 %d 行对不上 —— 跑一遍 gaz tidy 就理好了" % stale,
                            "序至失准"))

    # 整机、器件里点到的单位,名录里得有 —— 打错一个字,连线就连不上
    for sheet, cols in ((SHEET_COMP, ("Research Insti", "Factory")),
                        (SHEET_SEMI, ("Research Insti", "Factory"))):
        if sheet not in wb.sheetnames:
            continue
        w = wb[sheet]
        hh = _headers(w, 1)
        for r in range(2, w.max_row + 1):
            who = w.cell(row=r, column=hh["Product"]).value if "Product" in hh else ""
            book = _book_of(w.cell(row=r, column=hh["Remark"]).value
                            if "Remark" in hh else "")
            if book:
                who = "%s〔%s〕" % (who or "", book)
            miss = []
            for c in cols:
                if c not in hh:
                    continue
                for one in ALIAS_SPLIT.split(str(w.cell(row=r, column=hh[c]).value or "")):
                    one = _bare(one)
                    if one and one not in names and one not in miss:
                        miss.append(one)
            if miss:
                bad.append(("名录", "%s 第%d行 %s" % (sheet, r, who or ""),
                            "、".join(miss),
                            "%s·%s·%s" % (sheet, who or "", "、".join(miss))))
    return bad
