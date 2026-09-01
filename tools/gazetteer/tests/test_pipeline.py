#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整条流水线的回归测试:纪年折算 → Markdown → 抽取 → 回写 xlsx。

不依赖 pytest,直接跑:

    python3 tools/gazetteer/tests/test_pipeline.py

样张见 fixture/README.md,盯的是几处最容易张冠李戴的地方。
"""

import os
import re
import shutil
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", ".."))

from gazetteer import (bookmd, cndate, extract as EX, notes,  # noqa: E402
                       toxlsx, tsvio, vault)

FAILED = []

# geocode.js 里 CITY_FALLBACK 收了哪几个市 —— extract.OTHER_CITY 不能比它多,
# 多出来的市在图上没有落点,认出来反而会掉进上海的兜底里
GEOCODE_CITIES = set(re.findall(
    r"^  ([A-Z][a-zA-Z]+): \{ lat:",
    open(os.path.join(REPO, "src", "geocode.js"), encoding="utf-8").read(), re.M))


def check(cond, what):
    print(("  ✓ " if cond else "  ✗ ") + what)
    if not cond:
        FAILED.append(what)


def eq(got, want, what):
    check(got == want, "%s(得 %r)" % (what, got) if got != want else what)


# ---------------------------------------------------------------- 纪年

def test_dates():
    print("纪年折算")
    eq(cndate.parse("1958年3月"), "19580300", "1958年3月")
    eq(cndate.parse("一九五八年三月十日"), "19580310", "一九五八年三月十日")
    eq(cndate.parse("民国二十六年"), "19370000", "民国二十六年")
    eq(cndate.parse("1958年上半年"), "19580000", "季度写法退到年")
    eq(cndate.parse("19660601"), "19660601", "已是八位写法")
    eq(cndate.parse("没有日期"), None, "没有日期就留空")
    eq(cndate.fmt("19730400"), "1973.04", "给人看的写法")


# ---------------------------------------------------------------- 认名字

def test_names():
    print("单位名")
    eq(EX.unit_names("该机与上海无线电十三厂协作生产。"), ["上海无线电十三厂"], "连接词后头的名字")
    eq(EX.unit_names("由上海红星电讯厂和上海新华电器厂合并组建。"),
       ["上海红星电讯厂", "上海新华电器厂"], "并列的两家")
    eq(EX.unit_names("上海和平电器厂由某某厂改组。"), ["上海和平电器厂"], "名字里带「和」不被截断")
    eq(EX.unit_names("1966年3月改名为上海无线电十九厂。"), ["上海无线电十九厂"], "剥掉年月与动词")
    eq(EX.unit_names("厂房面积15600平方米。"), [], "厂房不是厂")
    eq(EX.unit_names("该所所长由王某担任,全所职工486人。"), [], "所长不是所")
    # 以下都是拿《北京工业志·电子志》第三章跑出来的漏子
    eq(EX.unit_names("交由国营738厂试制。"), ["国营738厂"], "厂名里的数字不是断点")
    eq(EX.unit_names("四机部15所为空军125工程研制108甲。"), ["四机部15所"],
       "「四机部」是部委正名,不是「四个机部」")
    eq(EX.unit_names("生产厂家有北京计算机外部设备三厂。"), ["北京计算机外部设备三厂"],
       "「设备」的「设」不是断点,「厂家」也不算收尾")
    eq(EX.unit_names("北京北优计算机系统有限公司承接。"), ["北京北优计算机系统有限公司"],
       "「系统」的「系」不是断点")
    eq(EX.unit_names("华北计算技术研究所和哈尔滨军事工程学院联合研制。"),
       ["华北计算技术研究所", "哈尔滨军事工程学院"], "「和」连着的两家不并成一家")
    # 志书讲协作、讲配套,动词花样比改名多得多。下面这些是《北京工业志·
    # 电子志》第三章跑出来的原样脏名字,111 家里有二十几家是这么来的
    eq(EX.unit_names("1972年，中国科学院自动化研究所与该厂协作研制。"),
       ["中国科学院自动化研究所"], "「自」不再把「自动化」腰斩")
    eq(EX.unit_names("样机交由中国科学院的508所试验。"), ["508所"],
       "「的」是断点,「508所」是名字")
    eq(EX.unit_names("中国科学院半导体研究所研制成功3AG管。"), ["中国科学院半导体研究所"],
       "「科学院」不是「学院」,别从长名里截出个「中国科学院」")
    eq(EX.unit_names("1990年完成了新疆水泥厂的过程控制改造。"), ["新疆水泥厂"], "剥掉「完成了」")
    eq(EX.unit_names("1988年通过对DEC公司微机的仿制。"), ["DEC公司"], "「通过对」两层都剥掉")
    # 承接的那一头是拿机器去用的人家 —— 正是要留的,不是要丢的
    eq(EX.unit_names("该所承接了市旅游汽车公司的调度系统。"), ["市旅游汽车公司"],
       "「承接了」后头切开,用户单位留下来")
    eq(EX.unit_names("四机部6所承接了大庆石油化工总厂的过程控制系统。"),
       ["四机部6所", "大庆石油化工总厂"], "研制方与用户两家都认出来")
    eq(EX.clean_unit_name("北京华海新技术开发公司"), "北京华海新技术开发公司",
       "「开发」不当动词切 —— 它在这儿是名号的一截")
    eq(EX.unit_names("该校电子厂生产微机。"), [], "「该校电子厂」截出来的「校电子厂」不算名字")
    eq(EX.clean_unit_name("第一个通过机电部4所"), "机电部4所", "「通过」后头切开,所名留下")
    eq(EX.clean_unit_name("第一机床厂"), "第一机床厂", "真的「第一…厂」照留")
    eq(EX.unit_names("首钢电子公司（当时名为自动化研究所）开发建成。"),
       ["首钢电子公司", "自动化研究所"], "「自」当连接词剥得,「自动化」剥不得")
    eq(EX.clean_unit_name("6所华胜计算机公司"), "华胜计算机公司", "「6所」是定语,不是名号")
    eq(EX.clean_unit_name("508所"), "508所", "可「508所」本身就是名号,剥不得")
    eq(EX.find_district(["厂址杨浦区平凉路1690号。"]), "杨浦", "区名不带前一个字")
    eq(EX.find_address(["上海元件十四厂,厂址在闸北区中华新路688号。"])[0], "中华新路688号",
       "地址照原表体例,不带区名")
    # 「集成电路」的「路」不是马路 —— 国营738厂的地址一度写着「自动导航专用的集成电路」
    eq(EX.find_address(["国营738厂研制自动导航专用的集成电路。"])[0], "", "集成电路不是地址")
    eq(EX.find_address(["该厂研制印刷线路板。"])[0], "", "印刷线路也不是")
    # 「厂址」明说了是地址,没门牌号也照抄 —— 志书写到哪一步,就录到哪一步。
    # 「厂址位于朝阳区深沟村」与「厂址在亚运村」是一个样子,不能一个收一个不收。
    eq(EX.find_address(["厂址在亚运村。"])[0], "亚运村", "厂址明说的,没门牌号也照录")
    eq(EX.find_address(["厂址位于朝阳区深沟村。"])[0], "深沟村", "村也是地名,不必有门牌号")
    # 「设」「迁至」什么都能领,那一类还是得有门牌号才作数
    eq(EX.find_address(["该厂设在亚运村。"])[0], "", "「设」领的,没门牌号不收")
    eq(EX.find_address(["该厂在亚运村。"])[0], "", "没有字样领着的,更不收")
    # 北京的写法:胡同、N条、甲N号 —— 只认上海的路与弄,整章的厂址都抓不着
    eq(EX.find_address(["厂址位于西城区德胜门外塔院胡同8号。"])[0], "德胜门外塔院胡同8号",
       "胡同")
    eq(EX.find_address(["厂址在朝阳门外二条67号。"])[0], "朝阳门外二条67号", "「N条」")
    eq(EX.find_address(["厂址位于石景山区古城北路甲4号。"])[0], "古城北路甲4号", "甲N号")
    eq(EX.find_address(["厂址位于宣武区福长街四条4号。"])[0], "福长街四条4号",
       "一条地址里两个街名,取到最后那个门牌号")
    # 「位子」是「位于」的形近误认 —— 不认它,就会抓成「子东环北路42号」
    eq(EX.find_address(["位子东环北路42号，由市电器厂改建。"])[0], "东环北路42号",
       "「位子」当「位于」认,错字不粘进地址")
    # 「作坊」的「坊」不是街坊
    eq(EX.find_address(["前身是1956年由32家私营手工业作坊合并而成的联合厂。"])[0], "",
       "「作坊」不是地址")


# ---------------------------------------------------------------- 整条流水线

def test_pipeline():
    print("转换稿")
    # 扫描件 → Markdown 归 zhiconv 管(见 Historian_Archive_Management)。
    # 这里定死一份它那种成色的稿子,验的是抽取,不是转换。
    md, _enc = bookmd.read_text(os.path.join(HERE, "fixture", "上海电子仪表工业志.md"))
    check("<!-- p.101 -->" in md and "<!-- p.247 -->" in md, "页码锚点在")
    check("### 第一章 半导体器件" in md, "标题层级在")
    check("1958年6月建立。1966年3月" in md, "正文没有断行")

    print("抽取")
    known = toxlsx.merge_known(os.path.join(REPO, "CN_Electronic_Industry.xlsx"),
                               os.path.join(REPO, "src", "geocode.js"))
    res = EX.extract(md, book="上海电子仪表工业志", known=known)
    by = {r["Unit"]: r for r in res["units"]}

    check("上海无线电十九厂" in by and "上海元件十四厂" in by and "上海计算机研究所" in by,
          "三家专条单位都抽到了")
    u = by["上海无线电十九厂"]
    eq(u["Start Date"], "19580600", "十九厂始建年")
    eq(u["End Date"], "", "十九厂没有终止年(撤销的是元件十四厂)")
    check("并入" not in u["Founder"], "他家并入本厂的话没混进 Founder")
    eq(u["Add."], "平凉路1690号", "十九厂门牌")
    eq(u["staff"], 1218.0, "职工总数")
    eq(u["profit"], 432.8, "实现利润")
    eq(u["Industry"], "半导体", "行业")
    check(u["Source"].endswith("p.101"), "出处回注到页")

    v = by["上海元件十四厂"]
    eq(v["End Date"], "19851200", "元件十四厂终止年")
    check("并入上海无线电十九厂" in v["Remark"], "终局归并记进备注")
    check("并入" not in v["Founder"], "终局归并不进 Founder")
    check(v["Founder"].startswith("上海红星电讯厂和上海新华电器厂合并"), "合并来历")

    w = by["上海计算机研究所"]
    eq(w["Start Date"], "19610500", "计算所始建年")
    check("19650000改名上海市计算技术研究所" in w["Founder"], "中文数字纪年的改名")
    eq(w["district"], "徐汇", "区名")

    print("名称沿革")
    segs = [r for r in res["names"] if r["Unit"] == "上海计算机研究所"]
    eq(len(segs), 3, "计算所三段名称")
    check(all(r["Source"].startswith("上海电子仪表工业志") for r in segs),
          "沿革出处是志书页码,不是「据 Founder 列推定」")

    print("产品记录")
    semi = {(r["Product"], r["Factory"]): r for r in res["semi"]}
    check(("锗合金晶体管", "上海无线电十九厂") in semi, "器件挂到本厂")
    eq(semi[("锗合金晶体管", "上海无线电十九厂")]["Time"], "19590000", "器件投产年")
    eq(semi[("锗合金晶体管", "上海无线电十九厂")]["Personnel"], "周德昌", "人名不带职称")
    check(not any(r["Factory"] == "上海无线电十九厂" and r["Product"] == "可控硅整流器"
                  for r in res["semi"]), "别家的产品没串门")

    comp = {r["Product"]: r for r in res["comp"]}
    eq(comp["TQ-16型晶体管计算机"]["字长"], "48", "字长")
    eq(comp["TQ-16型晶体管计算机"]["Speed（次秒）"], "110000", "每秒11万次 → 110000")
    eq(comp["TQ-16型晶体管计算机"]["Personnel"], "虞浦帆、何育辽", "两位主持人")
    eq(comp["TQ-16型晶体管计算机"]["Factory"], "上海无线电十三厂", "协作厂")
    eq(comp["DJS-130型小型计算机"]["Factory"], "", "协作只算前一台机器")

    print("回写工作簿")
    tmp = tempfile.mkdtemp(prefix="gaz-test-")
    try:
        target = os.path.join(tmp, "wb.xlsx")
        shutil.copy2(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), target)
        import openpyxl
        before = toxlsx.units_sheet(openpyxl.load_workbook(target)).max_row
        picked = [dict(r, keep="y") for r in res["units"] if r["role"] == "专条"]
        rep = toxlsx.append(target, units=picked, semi=res["semi"], comp=res["comp"],
                            names=res["names"], backup=False, log=lambda *a: None)
        eq(rep["units"], 3, "追加 3 家单位")
        wb = openpyxl.load_workbook(target)
        ws = toxlsx.units_sheet(wb)
        eq(ws.max_row, before + 3, "名录表多了 3 行")
        row = [c for c in next(ws.iter_rows(min_row=before + 1, max_row=before + 1,
                                            values_only=True))]
        eq(row[0], "上海无线电十九厂", "A 列是单位名")
        eq(row[3], 19580600, "始建日期写成八位整数")
        eq(row[8], 1218, "统计块落在职工总数那一格")
        nh = toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES)
        last = [c for c in next(nh.iter_rows(min_row=nh.max_row, max_row=nh.max_row,
                                             values_only=True))]
        check(isinstance(last[2], str), "Name-History 的 From 照原表存成文本")

        print("Obsidian 笔记")
        vault = os.path.join(tmp, "vault")
        notes.write_vault(picked, vault, book="上海电子仪表工业志", book_note="上海电子仪表工业志",
                          log=lambda *a: None)
        with open(os.path.join(vault, "上海元件十四厂.md"), encoding="utf-8") as f:
            note = f.read()
        check("[[上海无线电十九厂]]" in note, "沿革里的单位连成 wikilink")
        check("始建: 1970.04" in note, "frontmatter 带始建年")
        check("原文佐证" in note, "原文照录")

        print("TSV 闸门")
        tsv = os.path.join(tmp, "u.tsv")
        tsvio.write(tsv, [dict(keep="?", Unit="甲"), dict(keep="y", Unit="乙")],
                    ["keep", "Unit"])
        eq([r["Unit"] for r in tsvio.kept(tsvio.read(tsv))], ["乙"], "只有 keep=y 的行放行")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)



def _read(p):
    with open(p, encoding="utf-8") as f:
        return f.read()


def _write(p, s):
    with open(p, "w", encoding="utf-8") as f:
        f.write(s)


def test_vault():
    """库 ←→ 工作簿的来回:改在 Obsidian 里,推回工作簿,再推回来不打架。"""
    print("库与工作簿的来回")
    import openpyxl
    tmp = tempfile.mkdtemp(prefix="gaz-vault-")
    try:
        xlsx = os.path.join(tmp, "wb.xlsx")
        shutil.copy2(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), xlsx)
        geo = os.path.join(REPO, "src", "geocode.js")
        vdir = os.path.join(tmp, "vault")
        quiet = lambda *a: None

        rep = vault.push(xlsx, vdir, geocode_js=geo, log=quiet)
        eq(rep["wrote"], len(toxlsx.read_units_full(xlsx)), "工作簿里有几家就写几则")
        check(os.path.exists(os.path.join(vdir, "_厂所索引.md")), "索引也写了")

        note = os.path.join(vdir, "上海微波设备研究所.md")
        fm = vault.parse_fm(vault.split_note(_read(note))[0])
        eq(fm["key"], "上海微波设备研究所", "key 是匹配用的钥匙")
        eq(fm["行业"], "研究所", "字段照工作簿")
        check(fm.get("推定坐标", "").startswith("31.27"), "geocode.js 的推定坐标只作参考")
        eq(fm["纬度"], "", "推定坐标不冒充数据,纬度仍空着")

        nh_before = len(toxlsx.read_name_history(xlsx))
        got = vault.collect(xlsx, vdir)
        eq(len(got["changes"]), 0, "刚推出来,没有该改的")
        # 沿革表里没有、而 Founder 链里有的,collect 会把它补出来 —— 那是它的本分。
        # 要紧的是一条也不能少。
        # 比「名称+启用年」,不比 Unit 那一格的写法:笔记按名录里的正名立,
        # 推回来时 Unit 会归到正名(「上海电子计算机厂」→「上海电子计算机厂（上无十三）」),
        # 那是归一,不是丢失
        nh_now = {(r["Name"], str(r["From"])) for r in toxlsx.read_name_history(xlsx)}
        if got["name_history"] is not None:
            back = {(r["Name"], str(r["From"])) for r in got["name_history"]}
            miss = sorted(nh_now - back)
            check(not miss, "推出来的沿革不该丢掉原表已有的行(丢了 %r)" % miss[:3])

        # 产品表混进名称沿革 —— 曾经真出过这个岔子
        big = os.path.join(vdir, "上海电子计算机厂.md")
        managed, _ = vault.split_managed(vault.split_note(_read(big))[1])
        check("整机" in managed, "整机表在 managed 段里")
        segs = vault.parse_nh_table(managed)
        # 从前拿「名字里有没有『计算机』」当判据,可「上海电子计算机厂」本就带着
        # 这三个字 —— 该问的是:有没有哪一段名称其实是产品表里的型号
        models = {toxlsx._bare(r["Product"]) for r in toxlsx.read_comp(xlsx)
                  if r.get("Product")}
        hit = [x["Name"] for x in segs if toxlsx._bare(x["Name"]) in models]
        check(not hit, "只读的产品表没被当成名称沿革(混进来的:%r)" % hit[:3])

        # 改三样:普通字段、宽写的日期、自填坐标;再改一行沿革的出处
        s0 = _read(note)
        # 城市这一格从前空着,如今表里填好了 —— 改动要作数就得换个值,
        # 别再赌工作簿里哪一格是空的
        newcity = "Nanjing"          # 只要与表里那格不同即可
        s0 = re.sub(r"(?m)^城市: .*$", "城市: " + newcity, s0)
        s0 = s0.replace("始建: 19501100", "始建: 1950年11月")
        s0 = s0.replace('纬度: ""', "纬度: 31.2701").replace('经度: ""', "经度: 121.3481")
        s0 = s0.replace("| 上海仪表铜厂 | 19610000 |  |  | 据 Founder 列推定,待核 |",
                        "| 上海仪表铜厂 | 19610000 | Copper Works | 1961年改称 | 仪表工业志 p.412 |")
        s0 = s0.replace("（这一段归你", "我的札记在此。（这一段归你")
        _write(note, s0)

        got = vault.collect(xlsx, vdir)
        eq(len(got["changes"]), 1, "认出一则改动")
        keys = {k for k, _o, _n in got["changes"][0]["shown"]}
        eq(keys, {"城市", "纬度", "经度"}, "只认出真改过的三样")
        check(got["name_history"] is not None, "沿革表也改了")

        rep = vault.pull(xlsx, vdir, log=quiet)
        eq(rep["units"], 1, "写回一家")
        check(len(toxlsx.read_name_history(xlsx)) >= nh_before, "沿革行只增不减")

        wb = openpyxl.load_workbook(xlsx)
        heads = [c.value for c in toxlsx.units_sheet(wb)[1]]
        check("Lat" in heads and "Lng" in heads, "Lat / Lng 两列按需添上")
        nh = [r for r in toxlsx.read_name_history(xlsx)
              if r["Unit"] == "上海微波设备研究所" and r["Name"] == "上海仪表铜厂"]
        eq(len(nh), 1, "改过的那一段还在")
        eq(nh[0]["Source"], "仪表工业志 p.412", "核实过的出处写回去了")
        eq(nh[0]["Name EN"], "Copper Works", "英文名不会因原表没这一列就丢掉")

        wb2 = toxlsx.units_sheet(openpyxl.load_workbook(xlsx))
        h = {c.value: i + 1 for i, c in enumerate(wb2[1]) if c.value}
        vals = [wb2.cell(row=r, column=h["Lat"]).value for r in range(3, wb2.max_row + 1)
                if wb2.cell(row=r, column=1).value == "上海元件五厂"]
        eq(vals, [None], "别家的推定坐标没被顺手写进表里")

        # pull 之后重新盖戳:再 push 不该把它当成未推的改动
        rep = vault.push(xlsx, vdir, geocode_js=geo, log=quiet)
        eq(rep["skipped"], [], "刚 pull 过,不该跳过")
        check("我的札记在此。" in _read(note), "自己写的札记,push 不动它")
        fm = vault.parse_fm(vault.split_note(_read(note))[0])
        eq(fm["城市"], newcity, "改动已在工作簿里,推回来还是它")

        # 真有未推的改动时,push 要拦住
        other = os.path.join(vdir, "上海元件五厂.md")
        _write(other, _read(other).replace("行业: 半导体", "行业: 电子计算机"))
        rep = vault.push(xlsx, vdir, geocode_js=geo, log=quiet)
        eq(rep["skipped"], ["上海元件五厂"], "未推的改动,push 拦住不盖")
        check("行业: 电子计算机" in _read(other), "拦住之后,库里的改动还在")
        rep = vault.push(xlsx, vdir, geocode_js=geo, force=True, log=quiet)
        eq(rep["skipped"], [], "--force 才照盖")
        check("行业: 半导体" in _read(other), "--force 之后以工作簿为准")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_book():
    """现成的转换稿:换个编码、换座城、没有页码锚点,照样走得通。"""
    print("现成的转换稿")
    import openpyxl
    src = os.path.join(HERE, "fixture", "北京转换稿-gb18030.md")

    text, enc = bookmd.read_text(src)
    eq(enc, "gb18030", "认出 GB18030,不是硬按 UTF-8 读")
    check("北京电子管厂" in text, "汉字没读成乱码")

    wrapped, share = bookmd.hard_wrapped(text)
    check(wrapped and share > 0.3, "认出照原书行宽硬断的稿子")
    check("北京市半导体\n器件研究所" in text, "厂名确实被断成了两截")
    flowed = bookmd.reflow_soft(text)
    check("北京市半导体器件研究所" in flowed, "接回段落之后厂名完整")
    got = EX.unit_names("北京市半导体\n器件研究所")
    check("北京市半导体器件研究所" not in got, "不接回就认不出整个厂名(所以才要接)")
    check(all("\n" not in n for n in got), "厂名里决不夹换行")

    picks = bookmd.detect_page_marks(flowed)
    eq(picks[0]["name"], "第N页", "认出「第N页」这种写法")
    check(picks[0]["ok"], "页码递增,像页码")
    normed, how, n = bookmd.normalize_page_marks(flowed)
    eq(how, "第N页", "采用它")
    eq(n, 3, "三处")
    check("<!-- p.55 -->" in normed and "<!-- p.57 -->" in normed, "归一成本工具的写法")

    # 脚注号不是页码 —— 这是「递增」这条判据要挡住的
    foot = "\n".join(["正文一。", "[1]", "正文二。", "[2]", "正文三。", "[1]",
                      "正文四。", "[3]", "正文五。", "[1]", "正文六。"])
    cands = [c for c in bookmd.detect_page_marks(foot) if c["ok"]]
    eq(cands, [], "[1][2][1] 这类脚注号没被当成页码")

    known = toxlsx.merge_known(os.path.join(REPO, "CN_Electronic_Industry.xlsx"),
                               os.path.join(REPO, "src", "geocode.js"))
    res = EX.extract(normed, book="北京工业志·电子志", known=known, city="Beijing")
    by = {r["Unit"]: r for r in res["units"]}
    eq(sorted(by), sorted(["北京电子管厂", "北京半导体器件二厂", "北京市半导体器件研究所"]),
       "三家专条单位")
    eq(by["北京电子管厂"]["City"], "Beijing", "City 列跟着 --city 走")
    eq(by["北京电子管厂"]["Add."], "酒仙桥路2号", "门牌")
    eq(by["北京电子管厂"]["staff"], 10286.0, "职工总数")
    eq(by["北京电子管厂"]["Source"], "北京工业志·电子志·p.55", "出处回注到页")
    eq(by["北京市半导体器件研究所"]["district"], "西城", "北京的区名也认得")
    # 动词照原文录:「划出组建」是分立,写成「合并」就成了另一回事
    check("划出组建" in by["北京半导体器件二厂"]["Founder"], "沿革里的动词照原文,不改字")

    # 没有页码时,出处退到篇章节
    nopage = re.sub(r"<!-- p\.\d+ -->", "", normed)
    res2 = EX.extract(nopage, book="北京工业志·电子志", known=known, city="Beijing")
    src2 = {r["Unit"]: r["Source"] for r in res2["units"]}
    check(src2["北京电子管厂"].startswith("北京工业志·电子志·"), "退到篇章节仍带书名")
    check("p.0" not in "".join(src2.values()), "决不写出 p.0 这种假页码")

    tmp = tempfile.mkdtemp(prefix="gaz-book-")
    try:
        out = os.path.join(tmp, "北京.xlsx")
        bookmd.write_xlsx(out, res, city="Beijing", book="北京工业志·电子志",
                          log=lambda *a: None)
        wb = openpyxl.load_workbook(out)
        eq(wb.sheetnames, ["待核", bookmd.UNITS_PREVIEW + "Beijing", toxlsx.SHEET_SEMI,
                           toxlsx.SHEET_COMP, toxlsx.SHEET_NAMES],
           "五张表,要核的那张排在头一个")
        ws = wb[bookmd.UNITS_PREVIEW + "Beijing"]
        eq([c.value for c in ws[1]][:8],
           [None, "Industry", "Product", "Start Date", "End Date", "Founder", "City", "Add."],
           "第一行表头与原表一致(A1 照原表留空)")
        eq(ws.cell(row=2, column=9).value, "职工总数", "第二行是统计块的表头")
        eq(ws.cell(row=3, column=4).value, 19561000, "日期写成八位整数")
        rv = wb["待核"]
        # 核的是名字,名字与据以判断的原文都摆在最左边
        eq(rv.cell(row=1, column=2).value, "单位", "待核表第二列就是单位名")
        eq(rv.cell(row=1, column=3).value, "别名", "别名紧挨着正名 —— 核的是名字")
        eq(rv.cell(row=1, column=6).value, "据以立论的原文", "原文也在近处,不用横拉")
        check(rv.max_row == len(res["units"]) + 1, "待核表一家一行")

        # 核过再读回来:「取否」写 y 的才算数,名字改成一样的并作一行
        rv.cell(row=2, column=1).value = "y"
        rv.cell(row=3, column=1).value = "y"
        rv.cell(row=2, column=2).value = "上海无线电十九厂"
        rv.cell(row=3, column=2).value = "上海无线电十九厂"
        rv.cell(row=4, column=1).value = ""
        wb.save(out)
        bundle, city2, seen = bookmd.read_review(out)
        eq(city2, "Beijing", "城市从「厂所名录-北京」这类表名上认")
        eq(len(bundle["units"]), 1, "两行改成同一个名字,并作一家")
        eq(seen["merged"], 1, "并了几行要报出来")
        eq(bundle["units"][0]["Unit"], "上海无线电十九厂", "并成的那一家用核过的名字")
        check("；" in bundle["units"][0]["Source"], "两处出处都留着")
        check(all(not r.get("evidence") for r in bundle["units"]), "原文只是核对用的,不进表")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_flat_heads():
    """别处转来的稿子:篇章节全压成 `##`,而且一个页码都没有。

    《北京工业志·电子志》2001 第三章就是这样。层级若只数井号,「第三章」
    「第一节」会被「一、」顶掉,而这书没有页码,出处全靠这条路径 —— 只剩
    「一、电子管计算机」等于没说,哪一章的一、都长这样。"""
    print("压平的标题 / 没有页码的书")
    text, enc = bookmd.read_text(os.path.join(HERE, "fixture", "北京-压平标题.md"))

    check(not bookmd.detect_page_marks(text),
          "认不出页码(这书本来就没有),不硬凑")

    blocks = EX.read_blocks(text)
    deep = [b for b in blocks if "晶体管" in "·".join(b["heads"])]
    check(bool(deep), "找得到「二、晶体管计算机」底下的正文")
    eq(deep[0]["heads"],
       ["第三章 电子计算机", "第一节 数字计算机", "二、晶体管计算机"],
       "三级标题同为 `##`,仍按字面认出章 → 节 → 一、")

    first = [b for b in blocks if "地区" in b["text"]][0]
    eq(first["heads"], ["第三章 电子计算机"], "概述段只归到章,不蹭下面的节")

    src_ = EX.make_source("北京工业志·电子志", None, "·".join(deep[0]["heads"]))
    for want in ("第三章", "第一节", "二、"):
        check(want in src_, "没页码时出处带上「%s」这一级" % want)

    res = EX.extract(text, book="北京工业志·电子志", city="Beijing", min_mentions=1)
    names = {r["Unit"] for r in res["units"]}
    check("北京计算机一厂" in names, "「北京计算 机一厂」中间的空格已去掉,厂名认全")
    check("机一厂" not in names, "没有截出「机一厂」这种鬼名字")
    check("国营738厂" in names, "「交由国营738厂试制」认得出国营738厂")


def test_fixes():
    """字形订正表:转换稿认错的字,一本书一张表。"""
    print("字形订正")
    tmp = tempfile.mkdtemp(prefix="gaz-fix-")
    try:
        tsv = os.path.join(tmp, "fixes.tsv")
        _write(tsv, "# 《北京工业志·电子志》2001\n"
                    "安做无线电厂\t安徽无线电厂\n"
                    "莫须有厂\t查无此厂\n")
        fixes = bookmd.load_fixes(tsv)
        eq(len(fixes), 2, "注解行不算一条")
        # 汉字之间的制表符是分栏用的,不能跟空格一起吃掉,不然两列并成一列
        eq(fixes[0], ("安做无线电厂", "安徽无线电厂"), "两列各归各的")
        text, ledger = bookmd.apply_fixes("由四机部6所、安做无线电厂联合研制。", fixes)
        check("安徽无线电厂" in text, "认错的字改了回来")
        eq([n for _a, _b, n in ledger], [1, 0], "每条改了几处都记着")
        check(bookmd.load_fixes(None) == [], "没给表就是没有,不报错")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_users():
    """用机的人家:记在整机的「用户」一列,不混进研制单位。

    先前一律记进 Factory,站点便据以推定「协作」—— 水泥厂成了计算机的共同
    研制单位。用户要留(机器去了哪儿,正是计算机怎么用开的),但不能算研制方。"""
    print("用户单位")
    md = ("## 一、控制机\n\n"
          "1988年，四机部6所承接了大庆石油化工总厂的过程控制系统，研制成功DJS-186小型计算机。\n"
          "1979年，北京控制机厂与北京市计算机技术研究所联合研制成功DJS-130小型计算机。\n")
    res = EX.extract(md, book="北京工业志·电子志", city="Beijing", min_mentions=1)
    by = {c["Product"]: c for c in res["comp"]}
    check("DJS-186小型计算机" in by, "承接工程那一台也记下来了")
    u = by["DJS-186小型计算机"]
    eq(u.get("用户"), "大庆石油化工总厂", "承接的那一头记进「用户」")
    check("大庆" not in u.get("Factory", ""), "决不混进研制单位 —— 否则要画成协作")
    check("四机部6所" in u.get("Factory", ""), "研制方还在原处")
    v = by["DJS-130小型计算机"]
    eq(v.get("用户", ""), "", "真的协作研制,没有用户")
    check("北京市计算机技术研究所" in (v.get("Research Insti", "") + v.get("Factory", "")),
          "真的协作单位照旧记下,协作连线不受影响")


def test_rename_subject():
    """一句话点到好几家时,「改名为」说的是主语那一家。"""
    print("改的是谁的名")
    md = ("## 一、控制机\n\n"
          "1979年，北京控制机厂为兰州炼油厂研制成功过程控制系统，同年改名为北京自动化设备厂。\n"
          "1980年，北京电器科学研究院参加该项目的设计。\n")
    res = EX.extract(md, book="试", city="Beijing", min_mentions=1)
    got = [(n["Unit"], n["Name"]) for n in res["names"]]
    eq(got, [("北京控制机厂", "北京自动化设备厂")], "只有主语那一家改了名")
    # 志书行文主语在前,「为兰州炼油厂」夹在主语与动词之间 —— 认第一家,不认最后一家
    check(EX.renames_this("北京控制机厂为兰州炼油厂研制系统，同年改名为甲厂。", 22, "北京控制机厂"),
          "主语那一家认得出")
    check(not EX.renames_this("北京控制机厂为兰州炼油厂研制系统，同年改名为甲厂。", 22, "兰州炼油厂"),
          "顺带提到的用户不算")
    check(EX.renames_this("该厂1966年改名为上海无线电十九厂。", 8, "上海元件十四厂"),
          "前头一家也没点到,那就是它")


def test_models():
    """型号常是句子的主语,从动词后头取不着。

    「104机的仿制与103机同时进行」—— find_products 一律从「试制/研制」之后取,
    这一句一台也取不着。按型号的样子再扫一遍,宁滥勿缺:漏掉一台就永远没有了,
    多认一个错的,核对时一眼划掉。"""
    print("型号")
    eq(EX.find_models("104机的仿制与103机同时进行。"), ["104机", "103机"], "主语位置上的型号")
    eq(EX.find_models("该机共有机柜32个，字长39位，内存4K，共生产38台。"), [],
       "「32个」「39位」「38台」都不是型号")
    eq(EX.find_models("其性能指标超过GB1962-82《声频功率放大器》的要求。"), [],
       "GB 开头是国标号,从来不是产品")
    eq(EX.model_key("103型通用数字计算机"), EX.model_key("103机"), "同一台机器,钥匙相同")
    # 「109乙计算机」跟的是「计算机」不是「机」—— 只认「机」「型」两个字,
    # 这一台连钥匙都配不出,跟表里旧的那条认不到一处去
    eq(EX.model_key("109乙计算机"), "109乙", "数字型号后头跟「计算机」也算")
    eq(EX.model_key("109乙晶体管计算机"), EX.model_key("109乙计算机"),
       "「109乙计算机」与「109乙晶体管计算机」是一台")
    for junk in ("32个", "39位", "38台", "1958年"):
        eq(EX.model_key(junk), "", "「%s」配不出钥匙" % junk)
    check(EX.model_key("DJS-130小型计算机") != EX.model_key("DJS2"), "不同型号,钥匙不同")

    md = ("## 第三章 电子计算机\n\n## 一、电子管计算机\n\n"
          "1958年，中国科学院计算技术研究所仿M-3试制103型通用数字计算机。\n"
          "104机的仿制与103机同时进行。104机比103机大得多，共有机柜32个。\n")
    res = EX.extract(md, book="北京工业志·电子志", city="Beijing", min_mentions=1)
    prod = {c["Product"] for c in res["comp"]}
    check("104机" in prod, "没人说谁造的,也照样进产品名录")
    check(not any(EX.model_key(p) == "103" and p != "103型通用数字计算机" for p in prod),
          "「103机」与「103型通用数字计算机」是一台,不重复登记")
    orphan = [c for c in res["comp"] if c["Product"] == "104机"][0]
    eq(orphan.get("Factory", ""), "", "研制单位空着就是空着,不硬派给谁")
    check("未详" in orphan["Remark"], "备注里写明研制单位未详")
    check(orphan["Remark"].endswith("一、电子管计算机"), "出处照旧回注到篇章节")


def test_output():
    """产量:「至1960年,共生产38台」。动词得在数字前头,不然满篇都是产量。"""
    print("产量")
    eq(EX.find_output("至1960年，共生产38台。")[0], 38, "共生产38台")
    eq(EX.find_output("1988年，年产微机1.2万台。")[0], 12000, "「1.2万台」折成个位")
    eq(EX.find_output("全机共用800多只电子管。")[0], 0, "电子管的只数不是产量")
    eq(EX.find_output("配套外部设备有磁鼓4台；磁带机5台。")[0], 0, "配套设备的台数不是产量")
    eq(EX.find_output("机房占地面积约40平方米。")[0], 0, "平方米不是产量")
    eq(EX.find_output("共生产12万只。")[0], 120000, "器件论只,整机论台,都要认")
    eq(EX.find_output("共生产集成电路120万块。")[0], 1200000, "集成电路论块")

    # 器件也分研制与生产:半导体所研制、元件厂投产。原表只有一列 Factory,
    # 研究所便也写作了厂
    md4 = ("## 第二章 半导体器件\n\n## 一、晶体管\n\n"
           "1965年，中国科学院半导体研究所研制成功3AG型锗晶体管。\n"
           "1968年，上海元件五厂投产3DG型硅晶体管，共生产12万只。\n")
    r4 = EX.extract(md4, book="试", city="Shanghai", min_mentions=1)
    by4 = {c["Product"]: c for c in r4["semi"]}
    eq(by4["3AG型锗晶体管"].get("Research Insti"), "中国科学院半导体研究所", "研究所记进研制单位")
    eq(by4["3AG型锗晶体管"].get("Factory", ""), "", "研究所不写作厂")
    eq(by4["3DG型硅晶体管"].get("Factory"), "上海元件五厂", "厂还是记进生产单位")
    eq(by4["3DG型硅晶体管"].get("产量"), "120000", "12万只")
    check(all(c["Product"] not in ("3AG型锗晶体管", "3DG型硅晶体管") for c in r4["comp"]),
          "半导体器件不该混进整机表 —— 哪怕这一章标题里带「电子」")

    # 同一个产品名只归一次类,不会一次进器件表、一次进整机表
    md5 = ("## 第三章 电子计算机\n\n## 一、电子管计算机\n\n"
           "1958年，中国科学院计算技术研究所试制103型通用数字计算机。104机的仿制同时进行。\n\n"
           "## 二、器件\n\n104机所用元件由上海元件五厂生产。\n")
    r5 = EX.extract(md5, book="试", city="Beijing", min_mentions=1)
    both = {c["Product"] for c in r5["semi"]} & {c["Product"] for c in r5["comp"]}
    eq(sorted(both), [], "一个产品名不会同时出现在两张表里")

    # 产量常与型号隔着一两句,主语是「该机」或干脆省略
    md = ("## 第三章 电子计算机\n\n## 一、电子管计算机\n\n"
          "1958年，中国科学院计算技术研究所仿M-3试制103型通用数字计算机。"
          "该机共有三大机柜。至1960年，共生产38台。\n")
    res = EX.extract(md, book="北京工业志·电子志", city="Beijing", min_mentions=1)
    by = {c["Product"]: c for c in res["comp"]}
    eq(by["103型通用数字计算机"]["产量"], "38", "隔了两句也认得回来")
    check("共生产38台" in by["103型通用数字计算机"]["Remark"], "备注里留着原话,好核对")
    # 同一段里有两台机器,产量归产量前头最后点到的那一台 —— M-3 是仿的对象,不是它
    eq(by.get("M-3", {}).get("产量", ""), "", "一段两台机器,产量不派给另一台")


def test_edit_in_place():
    """核对是改字,不只是打勾:改过的字要照改过的样子读回来。"""
    print("直接改字")
    import datetime
    eq(bookmd.date_cell("1958"), "19580000", "只写年份,补成八位")
    eq(bookmd.date_cell("195803"), "19580300", "写到月,补成八位")
    eq(bookmd.date_cell("1958年3月"), "19580300", "中文写法照样认")
    eq(bookmd.date_cell(datetime.datetime(1958, 3, 1)), "19580301", "Excel 存成日期格也认")
    eq(bookmd.date_cell("待查"), "待查", "认不出就原样留着,不猜")

    tmp = tempfile.mkdtemp(prefix="gaz-edit-")
    try:
        import openpyxl
        md, _enc = bookmd.read_text(os.path.join(HERE, "fixture", "北京-压平标题.md"))
        res = EX.extract(md, book="试", city="Beijing", min_mentions=1)
        x = os.path.join(tmp, "核.xlsx")
        bookmd.write_xlsx(x, res, city="Beijing", log=lambda *a: None)

        wb = openpyxl.load_workbook(x)
        # 一打开就该停在「待核」上 —— 停在预览表上,第一眼看见的是一列光秃秃的
        # 单位名,连「取否」列都没有,不知道该往哪儿写 y
        eq(wb.sheetnames[0], "待核", "「待核」排在头一张")
        eq(wb.active.title, "待核", "打开就停在「待核」")
        eq([w.title for w in wb.worksheets if w.sheet_view.tabSelected], ["待核"],
           "选中的只有「待核」这一张")
        rv = wb["待核"]
        head = [c.value for c in rv[1]]
        rv.cell(row=2, column=1).value = "y"
        rv.cell(row=2, column=head.index("单位") + 1).value = "安徽无线电厂"
        rv.cell(row=2, column=head.index("行业") + 1).value = "半导体"
        rv.cell(row=2, column=head.index("始建") + 1).value = 1958
        wb.save(x)

        bundle, city, seen = bookmd.read_review(x)
        eq(len(bundle["units"]), 1, "只取点了头的那一行")
        u = bundle["units"][0]
        eq(u["Unit"], "安徽无线电厂", "改过的厂名照改过的读回来")
        eq(u["Industry"], "半导体", "改过的行业照改过的读回来")
        eq(u["Start Date"], "19580000", "手填的年份补成八位,不会写成 1958")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_aliases():
    """一家单位、一台机器同时有好几个名号,不必挑一个 —— 全留着,全显示。"""
    print("别名")
    A = EX.find_aliases
    eq(A("中国科学院计算技术研究所（简称中科院计算所）成立。", "中国科学院计算技术研究所"),
       ["中科院计算所"], "「简称」后头那个")
    eq(A("仿M-3试制103型通用数字计算机（亦称DJS-1）。", "103型通用数字计算机"),
       ["DJS-1"], "「亦称」后头那个")
    eq(A("华北计算技术研究所（四机部15所）和哈军工联合研制。", "华北计算技术研究所"),
       ["四机部15所"], "没有「简称」二字,紧跟着的括号里也认")
    eq(A("北京崇文电子仪器厂（北京计算机五厂前身）。", "北京崇文电子仪器厂"),
       [], "「…前身」是注解,不是别名")
    eq(A("甲厂与乙厂协作，乙厂简称乙。", "甲厂"), [], "别名只跟紧挨着的正名走")

    # 别名与沿革是两回事:「四机部15所」是同时并行的另一个名号,「后改名为…」
    # 说的是后来改叫什么 —— 那有年份可系,归名称沿革表,不该混进别名
    who = "北京崇文电子仪器厂"
    for paren in ("后改名为北京计算机五厂", "改名为北京计算机五厂",
                  "1985年改名为北京计算机五厂", "后称北京计算机五厂",
                  "今北京计算机五厂", "原北京无线电三厂", "北京计算机五厂前身", "兼营"):
        eq(A("%s（%s）研制。" % (who, paren), who), [],
           "「%s」是沿革或注解,不是别名" % paren)

    # 改名那一年,取紧挨着动词的那个 —— 不是整句的年份
    md3 = ("## 一、控制机\n\n"
           "1965年，北京崇文电子仪器厂（1985年改名为北京计算机五厂）研制成功107机。\n")
    r3 = EX.extract(md3, book="试", city="Beijing", min_mentions=1)
    eq([(n["Unit"], n["Name"], n["From"]) for n in r3["names"]],
       [("北京崇文电子仪器厂", "北京计算机五厂", "19850000")],
       "改名记在1985年,不是造机器的1965年")
    check(not any(u.get("别名") for u in r3["units"]), "这一条不进别名列")

    md = ("## 一、电子管计算机\n\n"
          "1956年，中国科学院计算技术研究所（简称中科院计算所）仿M-3试制"
          "103型通用数字计算机（亦称DJS-1）。\n")
    res = EX.extract(md, book="试", city="Beijing", min_mentions=1)
    by = {u["Unit"]: u for u in res["units"]}
    eq(by["中国科学院计算技术研究所"]["别名"], "中科院计算所", "单位的别名记在自己名下")
    eq({c["Product"]: c["别名"] for c in res["comp"]}.get("103型通用数字计算机"),
       "DJS-1", "机器的别名记在机器名下")

    # 并作一行时,被并掉的名字不能就这么没了
    rows = [{"Unit": "华北计算技术研究所", "别名": "四机部15所", "Source": "甲"},
            {"Unit": "华北计算技术研究所", "别名": "电子部15所、电子部第15所", "Source": "乙"}]
    got, n = bookmd.merge_by_name(rows)
    eq(n, 1, "两行并作一行")
    eq(set(got[0]["别名"].split("、")), {"四机部15所", "电子部15所", "电子部第15所"},
       "两边的别名都留着,一边有两个也不丢")

    # 三个别名一路走到工作簿:待核 → read_review → toxlsx → 名录表
    tmp = tempfile.mkdtemp(prefix="gaz-alias-")
    try:
        import openpyxl
        many = "试甲所、试乙所、试丙所"
        md2 = "## 一、机\n\n1959年，辽阳试验计算技术研究所研制成功机器23计算机。\n"
        r2 = EX.extract(md2, book="试", city="Beijing", min_mentions=1)
        bk = os.path.join(tmp, "待核.xlsx")
        bookmd.write_xlsx(bk, r2, city="Beijing", log=lambda *a: None)
        wb = openpyxl.load_workbook(bk)
        rv = wb["待核"]
        hd = [c.value for c in rv[1]]
        for i in range(2, rv.max_row + 1):
            if rv.cell(row=i, column=2).value == "辽阳试验计算技术研究所":
                rv.cell(row=i, column=1).value = "y"
                rv.cell(row=i, column=hd.index("别名") + 1).value = many
        wb.save(bk)

        bundle, _city, _seen = bookmd.read_review(bk)
        eq(bundle["units"][0]["别名"], many, "三个别名原样读回来")

        master = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), master)
        toxlsx.append(master, backup=False, **bundle)
        ws = openpyxl.load_workbook(master)[toxlsx.SHEET_UNITS]
        hh = {c.value: c.column for c in ws[1] if c.value}
        check("别名" in hh, "原表没有「别名」列,按需在表尾添上")
        hit = [ws.cell(row=i, column=hh["别名"]).value
               for i in range(3, ws.max_row + 1)
               if ws.cell(row=i, column=1).value == "辽阳试验计算技术研究所"]
        eq(hit, [many], "三个别名一并写进名录表(站点按顿号拆开,逐个都能认)")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_no_dup():
    """再跑一遍不该多出一份 —— 单位连别名一起比,产品与沿革按整条比。"""
    print("不重复登记")
    tmp = tempfile.mkdtemp(prefix="gaz-dup-")
    try:
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        bundle = dict(
            units=[{"Unit": "辽阳试验计算技术研究所", "别名": "试甲所", "City": "Beijing"}],
            comp=[{"Product": "试验机零号", "Factory": "辽阳试验计算技术研究所", "Time": "19591100"}],
            semi=[{"Product": "试验管零号", "Factory": "辽阳试验元件厂", "Time": "19650000"}],
            names=[{"Unit": "辽阳试验计算技术研究所", "Name": "试乙所", "From": "19820000"}])
        a = toxlsx.append(x, backup=False, **bundle)
        eq([a["units"], a["semi"], a["comp"], a["names"]], [1, 1, 1, 1], "第一遍四样各进一条")
        b = toxlsx.append(x, backup=False, **bundle)
        eq([b["units"], b["semi"], b["comp"], b["names"]], [0, 0, 0, 0], "第二遍一条也不进")
        eq(len(b["skipped"]), 4, "四条都记在跳过里")

        # 「四机部15所」是「华北计算技术研究所」的别名,不是另一家
        c = toxlsx.append(x, backup=False, units=[{"Unit": "试甲所", "City": "Beijing"}])
        eq(c["units"], 0, "拿别名当正名送进来,认得出是同一家")

        d = toxlsx.append(x, backup=False, allow_dup=True, **bundle)
        eq(d["units"], 1, "--allow-dup 时照收不误")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_drafts_default():
    """没设 GAZ_DRAFTS 时,稿子的去处是仓库里的 转换稿 —— 不该逼人先设环境变量。"""
    print("转换稿的默认去处")
    import subprocess
    gaz = os.path.join(HERE, "..", "gaz.py")
    env = dict(os.environ)
    env.pop("GAZ_DRAFTS", None)
    # 指到一个不存在的目录:错误话里要点出默认去处,免得人不知道稿子该往哪儿放
    out = subprocess.run([sys.executable, gaz, "volume", "--dir",
                          os.path.join(tempfile.gettempdir(), "没有这个目录"),
                          "查无此稿"],
                         capture_output=True, text=True, env=env, cwd=REPO)
    said = out.stdout + out.stderr
    check("转换稿" in said, "说得出默认去处是仓库里的 转换稿")
    check("找不到目录" in said, "说得出是哪个目录找不到")

    tmp = tempfile.mkdtemp(prefix="gaz-drafts-")
    try:
        open(os.path.join(tmp, "某某志 第九章.md"), "w",
             encoding="utf-8").write("## 一、机\n\n无事。\n")
        out = subprocess.run([sys.executable, gaz, "volume", "--dir", tmp,
                              "第九章", "Beijing"],
                             capture_output=True, text=True, env=env, cwd=REPO)
        check("--city" in (out.stdout + out.stderr),
              "把城市当成关键词写,提醒该用 --city")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_accepted():
    """看过、认下的毛病不再翻出来 —— 二十八条旧账压着,新伤就没人看得见。"""
    print("《已核》")
    tmp = tempfile.mkdtemp(prefix="gaz-acc-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)

        eq(toxlsx.load_accepted(x), set(), "还没记过,一条也没有")
        base = toxlsx.verify(x)
        _p, n = toxlsx.save_accepted(x, base)
        eq(n, len({(t[0], t[3]) for t in base}), "眼下这些一律记下")

        seen = toxlsx.load_accepted(x)
        fresh = [t for t in toxlsx.verify(x) if (t[0], t[3]) not in seen]
        eq(fresh, [], "再验一遍,一条新的也没有")

        # 新伤照报不误
        wb = openpyxl.load_workbook(x)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS)
        h = toxlsx._headers(ws, 2)
        ws.cell(row=3, column=h["Start Date"]).value = 1961
        wb.save(x)
        fresh = [t for t in toxlsx.verify(x) if (t[0], t[3]) not in seen]
        eq(len(fresh), 1, "手改坏的那一处,认过的旧账压不住它")
        eq(fresh[0][0], "日期", "报的是日期这一类")

        # 行号会挪,记号不含行号 —— 插一行不该把旧账变成新账
        ws.cell(row=3, column=h["Start Date"]).value = None
        ws.insert_rows(3)
        ws.cell(row=3, column=1).value = "试验插进来的一行"
        ws.cell(row=3, column=h["Source"]).value = "试·一页"
        wb.save(x)
        fresh = [t for t in toxlsx.verify(x) if (t[0], t[3]) not in seen]
        eq(fresh, [], "插一行,底下行号全挪,认过的还是认过的")

        # 改好一条,--accept 重写时它自己掉出去。
        # 挑一条眼下真报着「没出处」的来补 —— 哪一家缺出处会随补录而变,不认死行号
        blank = next(r for r in range(3, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value
                     and not ws.cell(row=r, column=h["Source"]).value)
        ws.cell(row=blank, column=h["Source"]).value = "试·补的出处"
        wb.save(x)
        _p, n2 = toxlsx.save_accepted(x, toxlsx.verify(x))
        check(n2 < n, "改好的那条不再记进《已核》,名单不会越积越长")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_version():
    """gaz version 要答得出「手里这份是什么时候的」—— 拿旧版跑,少的那道关卡不吭声。"""
    print("版本自报")
    import subprocess
    out = subprocess.run([sys.executable, os.path.join(HERE, "..", "gaz.py"), "version"],
                         capture_output=True, text=True, cwd=REPO)
    said = out.stdout + out.stderr
    for want in ("版本", "分支", "认得的命令"):
        check(want in said, "报得出「%s」" % want)
    # 命令列表是拿来对照的:少了哪一个,手里那份就是旧的
    for cmd in ("verify", "tidy", "volume", "version"):
        check(cmd in said, "命令列表里有 %s" % cmd)


def test_run_out_encoding():
    """外部命令的输出按 UTF-8 解,不跟本机的编码走。

    简体中文 Windows 的本地编码是 GBK。subprocess 的 text=True 按本地编码解,
    而 git 吐出来的提交说明是 UTF-8 —— 里头一个 GBK 认不得的字节,读输出那个
    线程就当场炸掉,stdout 变成 None,底下 .strip() 报一句莫名其妙的
    AttributeError。真绊过人:gaz version 在中文提交说明上整个跑不动。"""
    print("外部命令的输出怎么解")
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "gaz_mod", os.path.join(HERE, "..", "gaz.py"))
    gaz = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(gaz)

    # 「」『』—— 这几个字的 UTF-8 字节 GBK 解不开,正是绊住的那一类
    zh = "判重加一道「抽字缩的简称」"
    got = gaz.run_out(sys.executable, "-c",
                      "import sys;sys.stdout.buffer.write(%r)" % zh.encode("utf-8"))
    eq(got, zh, "中文输出原样拿得回来")
    eq(gaz.run_out(sys.executable, "-c", "raise SystemExit(3)"), None,
       "命令返回非零,报 None,不是半截输出")
    eq(gaz.run_out("这个命令根本不存在-xyz"), None, "命令不存在,报 None,不抛异常")


def test_model_dash():
    """型号里的连字符被认成汉字「一」—— 一篇 117 处,不改就跟总表对不上。"""
    print("型号里的「一」")
    got, n = bookmd.fix_model_dash("TQ一16中型通用计算机,DJS一131,X一2型,JDK一331型")
    eq(n, 4, "四处都改了")
    eq(got, "TQ-16中型通用计算机,DJS-131,X-2型,JDK-331型", "改成连字符")
    # 中文里的「一」照旧
    for keep in ("第一台103型电子管数字计算机", "一九六五年", "全国第一家", "一五时期"):
        eq(bookmd.fix_model_dash(keep)[1], 0, "「%s」里的「一」不动" % keep)
    eq(bookmd.fix_model_dash("上海一厂")[1], 0, "汉字之间的「一」不动")


def test_product_attributive():
    """机器名前头的定语,不该连着录进来。"""
    print("机器名里的定语")
    # find_products 收的是**句子的列表**,不是一整串 —— 传字符串它会按字拆开,
    # 结果永远是空的,断言便句句落空
    P = lambda *sents: EX.find_products(list(sents))
    got = P("华东计算技术研究所从1969年开始研制运载火箭的箭载计算机，"
            "于1972年研制出KS-1箭载计算机。")
    check("箭载计算机" in got, "「运载火箭的箭载计算机」→ 箭载计算机(得 %r)" % got)
    check(not any("的" in x for x in got), "剥干净了,没留下带「的」的")

    got = P("1975年，该厂研制并投产的JS系列工业控制机。")
    check(got and not any("的" in x for x in got), "「并投产的」剥掉(得 %r)" % got)
    got = P("1980年，研制成功运算速度达100万次的大型计算机。")
    check(got and not any("的" in x for x in got),
          "「运算速度达100万次的」剥掉(得 %r)" % got)


def test_review_name_sheet():
    """待核那张沿革表:表头要说得死,表头行末那句注不许当成一列数据读回来。"""
    print("待核沿革表的表头")
    tmp = tempfile.mkdtemp(prefix="gaz-nh-")
    try:
        import openpyxl
        md, _enc = bookmd.read_text(os.path.join(HERE, "fixture", "上海电子仪表工业志.md"))
        res = EX.extract(md, book="试")
        x = os.path.join(tmp, "核.xlsx")
        bookmd.write_xlsx(x, res, city="Shanghai", log=lambda *a: None)

        wb = openpyxl.load_workbook(x)
        nh = toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES)
        head = [nh.cell(row=1, column=c).value for c in range(1, 8)]
        eq(head, ["取否", "序", "单位(今名)", "当时名称", "自哪年起", "Remark", "Source"],
           "表头写成中文,「Unit」不再看着像「这一行这家叫什么」")
        note = str(nh.cell(row=1, column=9).value or "")
        check(note.startswith("↑") and "不是它当时的名字" in note, "表头行末缀着念法")

        nh.cell(row=2, column=1).value = "y"
        want = (nh.cell(row=2, column=3).value, nh.cell(row=2, column=4).value)
        wb.save(x)
        bundle, _city, _seen = bookmd.read_review(x)
        got = bundle["names"][0]
        eq(sorted(got.keys()), ["From", "Name", "Remark", "Source", "Unit"],
           "读回来还是原来那几个字段名 —— 那句注没混成一列")
        eq((got["Unit"], got["Name"]), want, "中文表头照样对得上原字段")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_rename_verbs():
    """《北京工业志》第十八节那一段,一句话里三处漏 —— 逐条钉住。"""
    print("改名的几种说法")
    SENT = ("北京显像管总厂前身是北京呢绒服装厂，1970年改为北京市半导体器件五厂，"
            "生产低频大功率管。1972年11月，改名为北京显像管厂。"
            "1990年，北京半导体设备一厂并入北京显像管厂，改名为北京显像管总厂。")

    def nh(md, unit):
        r = EX.extract(md, book="试", city="Beijing", min_mentions=1)
        return [(n["Name"], n["From"]) for n in r["names"] if n["Unit"] == unit]

    got = nh("## 第十八节北京显像管总厂\n\n" + SENT + "\n", "北京显像管总厂")
    check(("北京市半导体器件五厂", "19700000") in got,
          "「1970年改为…」—— 改为也是改名的说法")
    check(("北京显像管厂", "19721100") in got, "同一节里没写主语的那一句,归本节这家")
    check(("北京呢绒服装厂", "19700000") not in got,
          "呢绒服装厂不该系上 1970 —— 那是它被改掉那年,不是启用那年")

    # 「改为」后头不像单位名的,不算改名
    for tail in ("民用产品", "全民所有制", "三班制"):
        r = EX.extract("## 一、厂\n\n1970年，北京某某厂改为%s。\n" % tail,
                       book="试", city="Beijing", min_mentions=1)
        eq([n["Name"] for n in r["names"]], [], "「改为%s」不是改名" % tail)

    # 「A并入B,改名为C」:改的是 B
    got = nh("## 一、厂\n\n1990年，北京半导体设备一厂并入北京显像管厂，"
             "改名为北京显像管总厂。\n", "北京半导体设备一厂")
    eq(got, [], "被并进去的那家没有改名")
    got = nh("## 一、厂\n\n1990年，北京半导体设备一厂并入北京显像管厂，"
             "改名为北京显像管总厂。\n", "北京显像管厂")
    eq(got, [("北京显像管总厂", "19900000")], "改名记在承接的那一家名下")

    # 前身那一截,定语不能连着捕进来
    r = EX.extract("## 一、辽阳无线电仪器二厂\n\n该厂前身之一的辽阳电子仪器厂，"
                   "1959年开始生产测量仪器。\n", book="试", city="Beijing", min_mentions=1)
    f = [u.get("Founder", "") for u in r["units"] if u["Unit"] == "辽阳无线电仪器二厂"]
    check(f and "之一的" not in f[0] and "辽阳电子仪器厂" in f[0],
          "「前身之一的X」剥掉定语,只留 X(得 %r)" % (f[0] if f else None))
    r = EX.extract("## 一、辽阳无线电厂\n\n该厂前身是全市第一家专门生产收音机的工厂。\n",
                   book="试", city="Beijing", min_mentions=1)
    check(not any("收音机的工厂" in u.get("Founder", "") for u in r["units"]),
          "「全市第一家专门生产收音机的工厂」是描述,不是名字")

    # 前身的生年:句子里写明「建于/创建于X」的,以它为准,不拿本厂的始建年顶替
    def pre(md, who):
        r = EX.extract(md, book="试", city="Beijing", min_mentions=1)
        return [(n["Name"], n["From"]) for n in r["names"] if n["Unit"] == who]
    eq(pre("## 一、辽阳配件三厂\n\n1990年建厂。辽阳配件三厂前身为辽阳塑料制品厂，建于1966年。\n",
           "辽阳配件三厂"), [("辽阳塑料制品厂", "19660000")],
       "前身系着它自己的生年 1966,不是本厂的 1990")

    # 有专条也不能一手遮天:句子里明写着别家作主语
    got = [(n["Unit"], n["Name"]) for n in EX.extract(
        "## 第十三节辽阳牡丹集团\n\n辽阳电视机厂前身为辽阳精密元件厂，创建于1965年。\n",
        book="试", city="Beijing", min_mentions=1)["names"]]
    check(all(u != "辽阳牡丹集团" for u, _n in got),
          "主语是电视机厂,这一段来历不记到本节那家名下(得 %r)" % got)

    # 一句话既讲前身又讲改名,两步都要记 —— 从前认下前身就 continue,改名那半句
    # 再也走不到,而志书正是这么写的
    r = EX.extract("## 一、辽阳无线电厂\n\n1958年建厂。辽阳无线电厂前身是辽阳电子厂，"
                   "1966年改名为辽阳仪器厂。\n", book="试", city="Beijing", min_mentions=1)
    f = [u.get("Founder", "") for u in r["units"] if u["Unit"] == "辽阳无线电厂"]
    check(f and "辽阳电子厂" in f[0] and "辽阳仪器厂" in f[0],
          "前身与改名同在一句,两步都进了沿革链(得 %r)" % (f[0] if f else None))


def test_diff_workbooks():
    """`.xlsx` 是二进制,git 只说「变了」—— 得有人说得出变了什么。"""
    print("比两份工作簿")
    tmp = tempfile.mkdtemp(prefix="gaz-diff-")
    try:
        import openpyxl
        a = os.path.join(tmp, "旧.xlsx")
        b = os.path.join(tmp, "新.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), a)
        shutil.copy(a, b)
        eq(toxlsx.diff_workbooks(a, b), {}, "一模一样的两份,报「没有分别」")

        wb = openpyxl.load_workbook(b)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS)
        h = toxlsx._headers(ws, 2)
        # 钥匙是去掉括号的样子:「上海电子计算机厂（上无十三）」→ 上海电子计算机厂
        who = toxlsx._bare(ws.cell(row=3, column=1).value)
        ws.cell(row=3, column=h["Source"]).value = "试·补的出处"
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = "辽阳试验新厂"
        toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES).delete_rows(2)
        wb.save(b)

        d = toxlsx.diff_workbooks(a, b)
        u = d[toxlsx.SHEET_UNITS]
        eq([x["Unit"] for x in u["added"]], ["辽阳试验新厂"], "添的那家认得出")
        eq(u["gone"], [], "没有谁凭空消失")
        chg = dict((k, dict((f[0], (f[1], f[2])) for f in v)) for k, v in u["changed"])
        check(who in chg and "Source" in chg[who], "改过的那一格,指名道姓")
        eq(chg[who]["Source"][1], "试·补的出处", "新值报得准")
        check(toxlsx.SHEET_NAMES in d and len(d[toxlsx.SHEET_NAMES]["gone"]) == 1,
              "删掉的那一段沿革,报作「消失」")

        # 「序」「至」是算出来的,不算数据变动:把 a 那份的这两列抹掉,
        # 再把 b 理一遍 —— 两份只差这两列
        shutil.copy(a, b)
        wb = openpyxl.load_workbook(a)
        nh = toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES)
        hn = toxlsx._headers(nh, 1)
        check("序" in hn and "至" in hn, "现表已经有「序」「至」两列")
        for i in range(2, nh.max_row + 1):
            for lab in ("序", "至"):
                nh.cell(row=i, column=hn[lab]).value = None
        wb.save(a)
        toxlsx.tidy_names(b)
        eq(toxlsx.diff_workbooks(a, b), {}, "只差算出来的那两列,数据上没有分别")
        d2 = toxlsx.diff_workbooks(a, b, skip_cols=())
        check(toxlsx.SHEET_NAMES in d2, "要连算出来的列一起比,那就报得出来")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_tidy_names():
    """沿革表要一眼看得出谁先谁后:同一单位挨在一处,按年份排,编上序号,补上「至」。"""
    print("理沿革表")
    tmp = tempfile.mkdtemp(prefix="gaz-tidy-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        wb = openpyxl.load_workbook(x)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES)
        h = toxlsx._headers(ws, 1)
        # 三段名称故意打乱着写进去,中间还夹一家别的
        r = ws.max_row + 1
        for unit, name, frm in (("辽阳试验所", "辽阳试验所", "19800000"),
                                ("别家厂", "别家旧名", "19600000"),
                                ("辽阳试验所", "辽阳甲厂", "19520000"),
                                ("辽阳试验所", "辽阳乙厂", "19650000")):
            ws.cell(row=r, column=h["Unit"]).value = unit
            ws.cell(row=r, column=h["Name"]).value = name
            ws.cell(row=r, column=h["From"]).value = frm
            r += 1
        wb.save(x)

        out = toxlsx.tidy_names(x)
        check(out["rows"] >= 4, "行数点得出来")

        ws = openpyxl.load_workbook(x)[toxlsx.SHEET_NAMES]
        h = toxlsx._headers(ws, 1)
        eq([ws.cell(row=1, column=c).value for c in range(1, 6)],
           ["序", "Unit", "Name", "From", "至"], "读起来顺的次序:序在最前,至挨着起")

        got = []
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=h["Unit"]).value == "辽阳试验所":
                got.append((ws.cell(row=i, column=h["序"]).value,
                            ws.cell(row=i, column=h["Name"]).value,
                            str(ws.cell(row=i, column=h["From"]).value),
                            ws.cell(row=i, column=h["至"]).value))
        eq([g[0] for g in got], [1, 2, 3], "同一单位的三段,序号 1、2、3")
        eq([g[1] for g in got], ["辽阳甲厂", "辽阳乙厂", "辽阳试验所"], "按年份先后排,不按写进去的顺序")
        eq([g[2] for g in got], ["19520000", "19650000", "19800000"], "起用年跟着名字走")
        eq([g[3] for g in got], ["19650000", "19800000", None],
           "「至」是下一段启用那年;末一段空着 —— 那是现在的名字")

        # 该空的那一格要真空:openpyxl 里 cell(..., value=None) 是「没给值」,清不掉
        idx = [i for i in range(2, ws.max_row + 1)
               if ws.cell(row=i, column=h["Unit"]).value == "辽阳试验所"][-1]
        check(ws.cell(row=idx, column=h["至"]).value is None, "末一段的「至」真的是空的,没留着旧字")

        # 手工改坏了,verify 报得出来,并指向 gaz tidy
        wb = openpyxl.load_workbook(x)
        toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES).cell(row=2, column=1).value = 9
        wb.save(x)
        why = [w for k, _wh, w, _kk in toxlsx.verify(x) if k == "沿革"]
        check(any("gaz tidy" in w for w in why), "序号改坏了,verify 指着 gaz tidy 说")

        toxlsx.tidy_names(x)
        why = [w for k, _wh, w, _kk in toxlsx.verify(x) if k == "沿革" and "tidy" in w]
        eq(why, [], "理过一遍就不再报")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_dups_near_names():
    """名字差一截的两家:「北大方正集团公司」与「北京北大方正集团公司」。
       判重只比名字相不相同,这种一个是另一个的一截,两边都过关。"""
    print("名字差一截的重复")
    tmp = tempfile.mkdtemp(prefix="gaz-near-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        wb = openpyxl.load_workbook(x)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS)
        r = ws.max_row + 1
        for nm in ("辽阳无线电器材厂", "辽阳无线电器材厂平谷分厂", "沈阳辽阳无线电器材厂"):
            ws.cell(row=r, column=1).value = nm
            r += 1
        wb.save(x)
        pairs = [k for _s, k, _rows, _w in toxlsx.report_dups(x)["similar"] if " ⊂ " in k]
        check(any("辽阳无线电器材厂 ⊂ 沈阳辽阳无线电器材厂" == k for k in pairs),
              "差一个前缀的,报出来")
        check(not any("平谷分厂" in k for k in pairs),
              "分厂本就是另一家,不报(得 %r)" % [k for k in pairs if "辽阳" in k])
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_dups_abbrev():
    """抽字缩的简称与全称各占一行:「上无十三」即「上海无线电十三厂」。
       既不同名,也不是掐头去尾差一截 —— 上面两道都拦不住,得照抽字查。"""
    print("简称与全称的重复")
    tmp = tempfile.mkdtemp(prefix="gaz-abbr-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        wb = openpyxl.load_workbook(x)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS)
        h = toxlsx._headers(ws, 2)
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = "辽阳电子仪器公司"
        ws.cell(row=r, column=h["别名"]).value = "辽无二厂、辽阳仪器"
        ws.cell(row=r + 1, column=1).value = "辽阳无线电二厂"
        ws.cell(row=r + 2, column=1).value = "抚顺辽阳仪器修配所"
        wb.save(x)
        pairs = [k for _s, k, _rows, _w in toxlsx.report_dups(x)["similar"] if " ⊂ " in k]
        check("辽无二厂 ⊂ 辽阳无线电二厂" in pairs,
              "抽字缩的简称,对得上全称就报出来")
        check("辽阳仪器 ⊂ 抚顺辽阳仪器修配所" not in pairs,
              "头一个字对不上的不报,不然满纸都是似是而非的对子")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_city_of():
    """志书按城市成书,书里点到的外地协作单位却不属本市 —— 名字冠了别的市名就照名字算。"""
    print("外地协作单位的城市")
    eq(EX.city_of("北京市半导体器件六厂", "Beijing"), "Beijing", "本市的照旧")
    eq(EX.city_of("唐山陡河发电总厂", "Beijing"), "Tangshan", "唐山的厂不算北京的")
    eq(EX.city_of("哈尔滨军事工程学院", "Beijing"), "Harbin", "长市名先比,不叫「哈」字截和")
    eq(EX.city_of("上海无线电七厂", "Beijing"), "Shanghai", "北京志里点到的上海厂归上海")
    eq(EX.city_of("华东计算技术研究所", "Shanghai"), "Shanghai", "名字不冠市名的,跟着志书走")
    for _zh, en in EX.OTHER_CITY.items():
        check(en in GEOCODE_CITIES,
              "%s 在 geocode.js 里有落点(没有就会掉进上海的兜底)" % en)


def test_verify():
    """手改之后的验一验:只报,一个格子也不动 —— 也不能漏报。"""
    print("手改之后验一验")
    tmp = tempfile.mkdtemp(prefix="gaz-vfy-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        toxlsx.append(x, backup=False, units=[
            {"Unit": "辽阳试验计算技术研究所", "City": "Beijing", "Source": "试·一页"}])

        def kinds(where_has):
            return sorted({k for k, w, _, _kk in toxlsx.verify(x) if where_has in w})

        base = toxlsx.verify(x)
        n0 = len(base)
        check(all(len(t) == 4 for t in base), "每一条都带着「哪一类」与「记号」")

        wb = openpyxl.load_workbook(x)
        ws = toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS)
        h = toxlsx._headers(ws, 2)
        r = None
        for i in range(3, ws.max_row + 1):
            if str(ws.cell(row=i, column=1).value or "").strip() == "辽阳试验计算技术研究所":
                r = i
        check(r is not None, "找得到刚添的那一行")

        ws.cell(row=r, column=h["Start Date"]).value = 1958
        wb.save(x)
        eq(kinds("辽阳"), ["日期"], "年份写成 1958 而不是 19580000,报出来")

        ws.cell(row=r, column=h["Start Date"]).value = "19580000"
        wb.save(x)
        eq(kinds("辽阳"), [], "补成八位就不再报")

        # 坐标:两个一起填才算数,填了还得落在中国境内
        col = max(h.values()) + 1
        ws.cell(row=2, column=col).value = "Lat"
        ws.cell(row=2, column=col + 1).value = "Lng"
        ws.cell(row=r, column=col).value = 41.27
        wb.save(x)
        eq(kinds("辽阳"), ["坐标"], "经纬度只填了一半,报出来")

        ws.cell(row=r, column=col + 1).value = 123.17
        wb.save(x)
        eq(kinds("辽阳"), [], "两个都填上,不再报")

        ws.cell(row=r, column=col).value = 123.17
        ws.cell(row=r, column=col + 1).value = 41.27
        wb.save(x)
        eq(kinds("辽阳"), ["坐标"], "经纬度填反了,落到中国境外,报出来")

        ws.cell(row=r, column=col).value = None
        ws.cell(row=r, column=col + 1).value = None
        ws.cell(row=r, column=h["Source"]).value = ""
        wb.save(x)
        eq(kinds("辽阳"), ["出处"], "出处空着,报出来")

        # 整机里点到的单位,名录里得有 —— 别名也算数,不能因为写了简称就报
        ws.cell(row=r, column=h["Source"]).value = "试·一页"
        ws.cell(row=r, column=h["别名"]).value = "辽试所"
        wb.save(x)
        c = toxlsx.sheet_of(wb, toxlsx.SHEET_COMP)
        hh = toxlsx._headers(c, 1)
        rr = c.max_row + 1
        c.cell(row=rr, column=hh["Product"]).value = "试验机零号"
        c.cell(row=rr, column=hh["Factory"]).value = "辽试所"
        wb.save(x)
        eq(kinds("试验机零号"), [], "点的是别名,认得出是同一家,不报")

        c.cell(row=rr, column=hh["Factory"]).value = "辽阳试验计算技术研究所、查无此厂"
        wb.save(x)
        eq(kinds("试验机零号"), ["名录"], "同一格里混着一个查无此人的,报出来")
        why = [w for k, wh, w, _kk in toxlsx.verify(x) if "试验机零号" in wh][0]
        eq(why, "查无此厂", "只点名查不到的那个,认得出的不跟着一起报")

        c.cell(row=rr, column=hh["Factory"]).value = "辽阳试验计算技术研究所"
        ws.cell(row=r, column=h["Source"]).value = "试·一页"
        wb.save(x)
        eq(len(toxlsx.verify(x)), n0, "都改回去,报的条数回到原样")

        # 产品表没有出处列,书名混在备注里 —— 判断「市计算机技术研究所」
        # 是哪个市的,全靠它
        eq(toxlsx._book_of("产量据「生产10台」。北京工业志·电子志·第三章"),
           "北京工业志", "备注末尾那截出处里的书名")
        eq(toxlsx._book_of("协作:某某厂。"), "", "备注里没有出处就不硬认")
        c.cell(row=rr, column=hh["Remark"]).value = "上海电子仪表工业志·第一章"
        c.cell(row=rr, column=hh["Factory"]).value = "查无此厂"
        wb.save(x)
        wh = [wh for k, wh, w, _kk in toxlsx.verify(x) if "试验机零号" in wh][0]
        check("〔上海电子仪表工业志〕" in wh, "报的时候带上是哪本志抄来的")
        c.cell(row=rr, column=hh["Remark"]).value = None
        c.cell(row=rr, column=hh["Factory"]).value = "辽阳试验计算技术研究所"
        wb.save(x)

        # 沿革表:1900 年代的年份是「1966年（后改名…」被切开算出来的
        nh = toxlsx.sheet_of(wb, toxlsx.SHEET_NAMES)
        hn = toxlsx._headers(nh, 1)
        nr = nh.max_row + 1
        nh.cell(row=nr, column=hn["Unit"]).value = "辽阳试验计算技术研究所"
        nh.cell(row=nr, column=hn["Name"]).value = "辽阳试验机械所"
        nh.cell(row=nr, column=hn["From"]).value = "19060000"
        wb.save(x)
        eq(kinds("%s 第%d行" % (toxlsx.SHEET_NAMES, nr)), ["日期"], "1906 这样的年份,报出来")

        nh.cell(row=nr, column=hn["From"]).value = "19660000"
        wb.save(x)
        eq(kinds("%s 第%d行" % (toxlsx.SHEET_NAMES, nr)), [], "1966 是正常年份,不报")

        # 「甲厂 → 甲厂」是改名链的末一段:前头有别的名字才立得住
        nh.cell(row=nr, column=hn["Name"]).value = "辽阳试验计算技术研究所"
        wb.save(x)
        eq(kinds("%s 第%d行" % (toxlsx.SHEET_NAMES, nr)), ["沿革"],
           "孤零零一条「改名叫自己」,不载信息,报出来")

        nr2 = nh.max_row + 1
        nh.cell(row=nr2, column=hn["Unit"]).value = "辽阳试验计算技术研究所"
        nh.cell(row=nr2, column=hn["Name"]).value = "辽阳试验机械所"
        nh.cell(row=nr2, column=hn["From"]).value = "19600000"
        wb.save(x)
        eq(kinds("%s 第%d行" % (toxlsx.SHEET_NAMES, nr)), [],
           "前头有一段别的名字,末一段就立得住,不报")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_verify_knows_geocode_aliases():
    """geocode.js 里认下的别名,verify 不该再报「查无此人」。

    「上海市计算中心」即上海市计算技术研究所 —— 别名表里写着,图上那条线连得
    好好的。verify 从前只翻工作簿,于是报它一句查无此人,白让人去核一趟。"""
    print("判名字时也认 geocode.js 的别名")
    tmp = tempfile.mkdtemp(prefix="gaz-ali-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        geo = os.path.join(REPO, "src", "geocode.js")
        wb = openpyxl.load_workbook(x)
        w = toxlsx.sheet_of(wb, toxlsx.SHEET_COMP)
        hh = toxlsx._headers(w, 1)
        r = w.max_row + 1
        w.cell(row=r, column=hh["Product"]).value = "试·某型计算机"
        # 一个只在 geocode.js 里认得的简称,一个哪儿都没有的
        w.cell(row=r, column=hh["Research Insti"]).value = "上海市计算中心、并无此厂"
        wb.save(x)

        def missed(bad):
            return "、".join(why for kind, _w, why, _k in bad
                             if kind == "名录" and "试·某型计算机" in _w)

        bare = missed(toxlsx.verify(x))
        check("上海市计算中心" in bare, "不给别名表,简称会被报出来(得 %r)" % bare)
        withgeo = missed(toxlsx.verify(x, geocode_js=geo))
        check("上海市计算中心" not in withgeo,
              "给了别名表,简称不再报(得 %r)" % withgeo)
        check("并无此厂" in withgeo, "真查无此人的照报不误(得 %r)" % withgeo)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_old_review_workbook():
    """手里核了一半的待核本子还挂着英文标签,照样读得回来。

    标签改名那天,谁手上正核着一章,那一章的工夫不能白费 —— 读的一头新旧
    两种都认。城市也一样:从前是「Fact and Comp-北京」,如今是「厂所名录-北京」。"""
    print("旧待核本子还认不认")
    tmp = tempfile.mkdtemp(prefix="gaz-oldrev-")
    try:
        import openpyxl
        res = EX.extract("# 第一章 电子计算机\n\n"
                         "## 第一节 辽阳无线电厂\n\n"
                         "辽阳无线电厂，创建于1965年3月，厂址辽阳路25号。\n",
                         book="试验志", city="Beijing")
        rev = os.path.join(tmp, "试验志待核.xlsx")
        bookmd.write_xlsx(rev, res, city="Beijing", book="试验志", log=lambda *a: None)

        wb = openpyxl.load_workbook(rev)
        eq(wb.sheetnames[1], bookmd.UNITS_PREVIEW + "Beijing", "新本子挂中文标签")

        # 改回旧标签,当作是改名以前做的那一份
        wb[bookmd.UNITS_PREVIEW + "Beijing"].title = bookmd.UNITS_PREVIEW_OLD + "Beijing"
        for new_name, old_name in toxlsx.OLD_NAMES.items():
            if new_name in wb.sheetnames:
                wb[new_name].title = old_name
        # 核过一行:取否写 y
        rv = wb["待核"]
        h = {c.value: c.column for c in rv[1]}
        rv.cell(row=2, column=h["取否"]).value = "y"
        wb.save(rev)

        bundle, city, seen = bookmd.read_review(rev)
        eq(city, "Beijing", "城市从旧写法的表名上照样认得出")
        eq(len(bundle["units"]), 1, "核过的那一行读得回来")
        eq(bundle["units"][0].get("Unit"), "辽阳无线电厂", "读回来的是那一家")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_verify_founder_years():
    """沿革链里的年份也得查 —— 从前只查 Start/End 两栏。

    「1966年（后改名…」被切开算成 1906,写进哪一栏都是错的。可 verify 只盯
    始建、终止两栏,写在 Founder 链里的就一直躺着没人看见 —— 现表里躺着两个。"""
    print("沿革链里的年份")
    tmp = tempfile.mkdtemp(prefix="gaz-fy-")
    try:
        import openpyxl
        x = os.path.join(tmp, "总表.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), x)
        wb = openpyxl.load_workbook(x)
        ws = toxlsx.units_sheet(wb)
        h = toxlsx._headers(ws, 2)
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = "辽阳试验电子厂"
        ws.cell(row=r, column=h["Founder"]).value = "19030000改名辽阳无线电厂"
        ws.cell(row=r, column=h["Source"]).value = "试·一页"
        wb.save(x)
        hit = [w for kind, w, why, _k in toxlsx.verify(x)
               if kind == "日期" and "1903" in why]
        check(hit, "链子里那个 1903 报出来了")
        # 正经年份不报,不然满纸都是
        ws.cell(row=r, column=h["Founder"]).value = "19630000改名辽阳无线电厂"
        wb.save(x)
        hit = [w for kind, w, why, _k in toxlsx.verify(x)
               if kind == "日期" and "辽阳试验电子厂" in w]
        eq(hit, [], "1963 是正经年份,不报")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_old_sheet_name():
    """标签改叫「厂所名录」了,手里的旧工作簿还得读得动。

    从前叫「Fact and Comp-Shanghai」—— 那时只收上海,后来北京、天津的厂所
    也进了这一张,名字就不对了。改名是为了跟文档里说的「厂所表」对得上,
    可别人手里、git 历史里的旧本子还是旧名字,认死一个就打不开了。"""
    print("旧标签名还认不认")
    tmp = tempfile.mkdtemp(prefix="gaz-tab-")
    try:
        import openpyxl
        old = os.path.join(tmp, "旧本子.xlsx")
        shutil.copy(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), old)
        wb = openpyxl.load_workbook(old)
        toxlsx.sheet_of(wb, toxlsx.SHEET_UNITS).title = toxlsx.OLD_NAMES[toxlsx.SHEET_UNITS]
        wb.save(old)
        eq(openpyxl.load_workbook(old).sheetnames[0], toxlsx.OLD_NAMES[toxlsx.SHEET_UNITS],
           "这一份用的是旧标签名")

        n_new = len(toxlsx.read_units_full(os.path.join(REPO, "CN_Electronic_Industry.xlsx")))
        eq(len(toxlsx.read_units_full(old)), n_new, "旧标签名照样读得出全部厂所")
        check(toxlsx.verify(old), "旧本子也验得动")
        # 一新一旧对着比,不该当成「整张表都没了」
        d = toxlsx.diff_workbooks(os.path.join(REPO, "CN_Electronic_Industry.xlsx"), old)
        gone = d.get(toxlsx.SHEET_UNITS, {}).get("gone", [])
        eq(gone, [], "改名前后两份对着比,厂所不该凭空消失(得 %r)" % gone[:3])
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_later_rename():
    """「后更名为」没说是哪一年 —— 宁可不记年份,也不能安上一个。"""
    print("后更名为")
    # 窗口按字数硬切,「1966年，…（后」切出「6年，…」,认成了 1906 年
    eq(EX.near_date("1966年，北京开关厂平谷分厂（后更名为甲厂）。", 17), None,
       "隔着逗号与括号的那个年份,不算改名那年")
    eq(EX.near_date("北京崇文电子仪器厂（1985年改名为甲厂）。", 19), "19850000",
       "同一小句里写明的年份,照写明的取")

    def nh(md):
        r = EX.extract(md, book="试", city="Beijing", min_mentions=1)
        return [(n["Unit"], n["Name"], n["From"]) for n in r["names"]]

    eq(nh("## 一、机\n\n1966年，北京开关厂平谷分厂（后更名为北京控制机厂）接受成果。\n"),
       [], "「后更名为」不记进沿革表 —— 没有年份可系")
    eq(nh("## 一、机\n\n1965年，北京崇文电子仪器厂（1985年改名为北京计算机五厂）研制成功107机。\n"),
       [("北京崇文电子仪器厂", "北京计算机五厂", "19850000")], "写明年份的照旧")
    # 「中心」不在收尾字样之列,一家也认不出 —— 那就谁也别派,别派给同句的汽车公司
    eq(nh("## 一、机\n\n1984年6月，北京市计算机软件中心（后更名为北京计算机五厂）"
          "承接了市旅游汽车公司的任务。\n"),
       [], "主语认不出来时,不把名字派给同句里别的单位")

    # 改名本身没丢 —— 它进了 Founder 链,只是不带年份
    r = EX.extract("## 一、机\n\n1966年，北京开关厂平谷分厂（后更名为北京控制机厂）接受成果。\n",
                   book="试", city="Beijing", min_mentions=1)
    f = [u.get("Founder", "") for u in r["units"] if u.get("Founder")]
    check(any("改名北京控制机厂" in x for x in f), "改名记在 Founder 链里,不带年份")
    check(not any("1906" in x or "19060000" in x for x in f), "决不写出 1906 这样的年份")


def main():
    for fn in (test_dates, test_names, test_pipeline, test_vault,
               test_book, test_flat_heads, test_fixes, test_users,
               test_rename_subject, test_models, test_output,
               test_edit_in_place, test_aliases, test_no_dup,
               test_drafts_default, test_version, test_run_out_encoding,
               test_dups_near_names,
               test_dups_abbrev, test_city_of,
               test_model_dash, test_product_attributive,
               test_review_name_sheet, test_rename_verbs, test_diff_workbooks, test_tidy_names, test_verify, test_accepted,
               test_verify_knows_geocode_aliases,
               test_old_sheet_name, test_old_review_workbook,
               test_verify_founder_years,
               test_later_rename):
        fn()
    print()
    if FAILED:
        print("%d 项没通过:" % len(FAILED))
        for f in FAILED:
            print("   -", f)
        return 1
    print("全部通过。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
